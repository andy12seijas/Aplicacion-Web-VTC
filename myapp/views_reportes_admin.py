import re
from openpyxl.styles import Font
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import ContratoCliente, Instalacion, PerfilUsuario, Soporte, TasaCambio, User, Plan, Cuadrilla, VentaDirecta, Material, MovimientoInventario, InventarioGlobal
from django.http import JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.core.management import call_command
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models import Q
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, colors

VE_TZ = pytz.timezone('America/Caracas')

def convertir_a_datetime_aware(fecha_str):
    """
    Convierte "2026-05-01" (string) a un datetime object aware 
    en la zona horaria de Venezuela, a las 00:00:00 horas.
    """
    if not fecha_str:
        return None
    try:
        # 1. Parseamos el string a un objeto date
        #    (formato YYYY-MM-DD, que es el que viene de tu template)
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        # 2. Combinamos la fecha con la hora 00:00:00 y la "localizamos"
        #    en la zona horaria de Venezuela.
        fecha_aware = VE_TZ.localize(datetime.combine(fecha_obj, datetime.min.time()))
        return fecha_aware
    except ValueError:
        return None

def es_admin(user):
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name='Administrador').exists())


@login_required
@user_passes_test(es_admin)
def reportes_view(request):
    """Vista principal de reportes unificada"""
    from myapp.models import User, Plan, Cuadrilla, Material
    # En reportes_view, agrega:
    instaladores = User.objects.filter(groups__name='Instalador').distinct().order_by('first_name', 'username')
    vendedores = User.objects.filter(
            groups__name__in=['Vendedor', 'Supervisor', 'Instalador']
        ).distinct().order_by('first_name', 'username')
    planes = Plan.objects.filter(activo=True)
    cuadrillas = Cuadrilla.objects.filter(activo=True)
    materiales = Material.objects.filter(activo=True)
    
    context = {
        'vendedores': vendedores,
        'planes': planes,
        'cuadrillas': cuadrillas,
        'materiales': materiales,
         'instaladores': instaladores,
    }
    return render(request, 'Admin/reporte.html', context)


@login_required
@user_passes_test(es_admin)
def reporte_ventas_json(request):
    """API para obtener datos de ventas (incluye COMPLETADO y EN_PROCESO)"""
    
    import pytz
    
    tipo_reporte = request.GET.get('tipo', 'simple')
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    vendedor_id = request.GET.get('vendedor', '')
    plan_id = request.GET.get('plan', '')
    estado_filtro = request.GET.get('estado', '')  # NUEVO
    busqueda = request.GET.get('busqueda', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    
    # Zona horaria de Venezuela
    VE_TZ = pytz.timezone('America/Caracas')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    # ¡MUY IMPORTANTE para la fecha HASTA!
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    # Incluir COMPLETADO y EN_PROCESO
    ventas = ContratoCliente.objects.filter(estado__in=['COMPLETADO', 'EN_PROCESO'])
    
    # ===== FILTRAR POR ESTADO (NUEVO) =====
    if estado_filtro:
        ventas = ventas.filter(estado=estado_filtro)
    
    # ===== FILTRAR POR FECHA USANDO DATETIME AWARE =====
    if fecha_desde_aware and fecha_hasta_aware:
        ventas = ventas.filter(
            Q(estado='COMPLETADO', fecha_completado__gte=fecha_desde_aware, fecha_completado__lte=fecha_hasta_aware) |
            Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_desde_aware, fecha_creacion__lte=fecha_hasta_aware)
        )
    elif fecha_desde_aware:
        ventas = ventas.filter(
            Q(estado='COMPLETADO', fecha_completado__gte=fecha_desde_aware) |
            Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_desde_aware)
        )
    elif fecha_hasta_aware:
        ventas = ventas.filter(
            Q(estado='COMPLETADO', fecha_completado__lte=fecha_hasta_aware) |
            Q(estado='EN_PROCESO', fecha_creacion__lte=fecha_hasta_aware)
        )
    
    if vendedor_id:
        ventas = ventas.filter(creado_por_id=vendedor_id)
    if plan_id:
        ventas = ventas.filter(plan_contratado_id=plan_id)
    if busqueda:
        ventas = ventas.filter(
            Q(cliente_potencial__nombre__icontains=busqueda) |
            Q(cliente_potencial__apellido__icontains=busqueda) |
            Q(cliente_potencial__cedula__icontains=busqueda) |
            Q(customer_id__icontains=busqueda)
        )
    
    total_registros = ventas.count()
    paginator = Paginator(ventas, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    if tipo_reporte == 'simple':
        data_list = []
        for v in page_obj:
            # ===== CONVERTIR FECHAS A VENEZUELA =====
            if v.estado == 'COMPLETADO' and v.fecha_completado:
                fecha_ve = v.fecha_completado.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            else:
                fecha_ve = v.fecha_creacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            
            data_list.append({
                'id': v.id,
                'cliente': v.nombre_completo,
                'customer_id': v.customer_id or 'N/A',
                'ods': v.ods or 'N/A',
                'plan': v.plan_contratado.nombre,
                'fecha': fecha_str,
                'vendedor': v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                'estado': v.get_estado_display(),
            })
    else:
        data_list = []
        for v in page_obj:
            # ===== CONVERTIR FECHAS A VENEZUELA =====
            if v.estado == 'COMPLETADO' and v.fecha_completado:
                fecha_ve = v.fecha_completado.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M')
            else:
                fecha_ve = v.fecha_creacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M')
            
            data_list.append({
                'id': v.id,
                'cliente': v.nombre_completo,
                'cedula': v.cedula,
                'telefono': v.telefono_principal,
                'correo': v.correo_electronico,
                'direccion': v.direccion_detallada[:100] if v.direccion_detallada else 'N/A',
                'plan': v.plan_contratado.nombre,
                'fecha': fecha_str,
                'vendedor': v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                'customer_id': v.customer_id or 'N/A',
                'ods': v.ods or 'N/A',
                'atr': v.atr or 'N/A',
                'estado': v.get_estado_display(),
            })
    
    # Calcular estadísticas considerando el filtro de estado
    if estado_filtro:
        completados = ventas.filter(estado='COMPLETADO').count() if estado_filtro == 'COMPLETADO' else 0
        en_proceso = ventas.filter(estado='EN_PROCESO').count() if estado_filtro == 'EN_PROCESO' else 0
    else:
        completados = ventas.filter(estado='COMPLETADO').count()
        en_proceso = ventas.filter(estado='EN_PROCESO').count()
    
    estadisticas = {
        'total_ventas': total_registros,
        'completados': completados,
        'pendientes': en_proceso,
    }
    
    return JsonResponse({
        'data': data_list,
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })


@login_required
@user_passes_test(es_admin)
def reporte_instalaciones_json(request):
    """API para obtener datos de instalaciones"""
    
    import pytz
    
    tipo_reporte = request.GET.get('tipo', 'simple')
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    cuadrilla_id = request.GET.get('cuadrilla', '')
    estado = request.GET.get('estado', '')
    busqueda = request.GET.get('busqueda', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    
    # Zona horaria de Venezuela
    VE_TZ = pytz.timezone('America/Caracas')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    # ¡MUY IMPORTANTE para la fecha HASTA!
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    instalaciones = Instalacion.objects.select_related('asignacion__cuadrilla', 'asignacion__contrato', 'asignacion__venta_directa', 'modelo_modem')
    
    # ===== FILTRAR POR FECHA USANDO DATETIME AWARE =====
    if fecha_desde_aware:
        instalaciones = instalaciones.filter(fecha_instalacion__gte=fecha_desde_aware)
    if fecha_hasta_aware:
        instalaciones = instalaciones.filter(fecha_instalacion__lte=fecha_hasta_aware)
    if cuadrilla_id:
        instalaciones = instalaciones.filter(asignacion__cuadrilla_id=cuadrilla_id)
    if estado == 'completada':
        instalaciones = instalaciones.filter(completada=True)
    elif estado == 'pendiente':
        instalaciones = instalaciones.filter(completada=False)
    
    # Búsqueda
    if busqueda:
        ids_coincidentes = []
        for inst in instalaciones:
            nombre_cliente = inst.nombre_cliente
            cedula_cliente = inst.cedula_cliente
            customer_id = inst.customer_id
            
            if (busqueda.lower() in nombre_cliente.lower() or 
                busqueda.lower() in cedula_cliente.lower() or 
                (customer_id and busqueda.lower() in customer_id.lower())):
                ids_coincidentes.append(inst.id)
        
        instalaciones = instalaciones.filter(id__in=ids_coincidentes)
    
    total_registros = instalaciones.count()
    paginator = Paginator(instalaciones, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    if tipo_reporte == 'simple':
        data_list = []
        for inst in page_obj:
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if inst.fecha_instalacion:
                fecha_ve = inst.fecha_instalacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            else:
                fecha_str = 'No registrada'
            
            data_list.append({
                'id': inst.id,
                'cliente': inst.nombre_cliente,
                'customer_id': inst.customer_id,
                'ods': inst.orden_servicio,
                'fecha': fecha_str,
                'cuadrilla': inst.asignacion.cuadrilla.nombre if inst.asignacion and inst.asignacion.cuadrilla else 'N/A',
                'estado': 'Completada' if inst.completada else 'Pendiente',
            })
    else:
        data_list = []
        for inst in page_obj:
            direccion = "N/A"
            try:
                if inst.asignacion and inst.asignacion.contrato:
                    direccion = inst.asignacion.contrato.direccion_detallada or "N/A"
                elif inst.asignacion and inst.asignacion.venta_directa:
                    direccion = getattr(inst.asignacion.venta_directa, 'direccion', None) or "N/A"
            except:
                direccion = "N/A"
            
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if inst.fecha_instalacion:
                fecha_ve = inst.fecha_instalacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M')
            else:
                fecha_str = 'No registrada'
            
            data_list.append({
                'id': inst.id,
                'cliente': inst.nombre_cliente,
                'cedula': inst.cedula_cliente,
                'direccion': direccion[:100] if direccion else 'N/A',
                'plan': inst.plan,
                'cuadrilla': inst.asignacion.cuadrilla.nombre if inst.asignacion and inst.asignacion.cuadrilla else 'N/A',
                'fecha': fecha_str,
                'estado': 'Completada' if inst.completada else 'Pendiente',
                'modelo': inst.modelo_modem.nombre if inst.modelo_modem else 'N/A',
                'serial': inst.sn_modem or 'N/A',
                'metros': inst.metros_utilizados,
                'customer_id': inst.customer_id,
                'ods': inst.orden_servicio,
            })
    
    estadisticas = {
        'total_instalaciones': total_registros,
        'completadas': instalaciones.filter(completada=True).count(),
        'pendientes': instalaciones.filter(completada=False).count(),
    }
    
    return JsonResponse({
        'data': data_list,
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })


@login_required
@user_passes_test(es_admin)
def reporte_soportes_json(request):
    """API para obtener datos de soportes"""
    
    import pytz
    
    tipo_reporte = request.GET.get('tipo', 'simple')
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    tipo_soporte = request.GET.get('tipo_soporte', '')
    estado = request.GET.get('estado', '')
    busqueda = request.GET.get('busqueda', '')
    cuadrilla_id = request.GET.get('cuadrilla', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    
    # Zona horaria de Venezuela
    VE_TZ = pytz.timezone('America/Caracas')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    # ¡MUY IMPORTANTE para la fecha HASTA!
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    soportes = Soporte.objects.all()
    
    # ===== FILTRAR POR FECHA USANDO DATETIME AWARE =====
    if fecha_desde_aware:
        soportes = soportes.filter(fecha_hora_servicio__gte=fecha_desde_aware)
    if fecha_hasta_aware:
        soportes = soportes.filter(fecha_hora_servicio__lte=fecha_hasta_aware)
    if estado:
        soportes = soportes.filter(estado=estado)
    if cuadrilla_id:
        soportes = soportes.filter(cuadrilla_id=cuadrilla_id)
    if tipo_soporte:
        soportes = soportes.filter(asignacion__ticket__tipo_soporte=tipo_soporte)
    if busqueda:
        soportes = soportes.filter(
            Q(asignacion__ticket__nombre__icontains=busqueda) |
            Q(asignacion__ticket__apellido__icontains=busqueda) |
            Q(asignacion__ticket__cedula__icontains=busqueda) |
            Q(asignacion__ticket__ticket_padre__icontains=busqueda)
        )
    
    total_registros = soportes.count()
    paginator = Paginator(soportes, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    if tipo_reporte == 'simple':
        data_list = []
        for s in page_obj:
            try:
                cliente_nombre = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else "N/A"
                cedula = s.asignacion.ticket.cedula if s.asignacion and s.asignacion.ticket else "N/A"
                customer_id = s.asignacion.ticket.customer_id if s.asignacion and s.asignacion.ticket else "N/A"
                ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else "N/A"
            except:
                cliente_nombre = "N/A"
                cedula = "N/A"
                customer_id = "N/A"
                ticket_padre = "N/A"
            
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if s.fecha_hora_servicio:
                fecha_ve = s.fecha_hora_servicio.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            elif s.fecha_creacion:
                fecha_ve = s.fecha_creacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            else:
                fecha_str = 'N/A'
            
            data_list.append({
                'id': s.id,
                'cliente': cliente_nombre,
                'customer_id': customer_id,
                'ticket_padre': ticket_padre,
                'fecha': fecha_str,
                'cuadrilla': s.cuadrilla.nombre if s.cuadrilla else 'N/A',
                'estado': s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado,
            })
    else:
        data_list = []
        for s in page_obj:
            try:
                cliente_nombre = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else "N/A"
                cedula = s.asignacion.ticket.cedula if s.asignacion and s.asignacion.ticket else "N/A"
                ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else "N/A"
                tipo_display = s.asignacion.ticket.get_tipo_soporte_display() if s.asignacion and s.asignacion.ticket else "N/A"
            except:
                cliente_nombre = "N/A"
                cedula = "N/A"
                ticket_padre = "N/A"
                tipo_display = "N/A"
            
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if s.fecha_hora_servicio:
                fecha_ve = s.fecha_hora_servicio.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M')
            elif s.fecha_creacion:
                fecha_ve = s.fecha_creacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M')
            else:
                fecha_str = 'N/A'
            
            data_list.append({
                'id': s.id,
                'ticket_padre': ticket_padre,
                'cliente': cliente_nombre,
                'cedula': cedula,
                'tipo': tipo_display,
                'estado': s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado,
                'fecha': fecha_str,
                'falla': s.falla_encontrada[:100] if s.falla_encontrada else 'N/A',
                'solucion': s.solucion[:100] if s.solucion else 'N/A',
                'cuadrilla': s.cuadrilla.nombre if s.cuadrilla else 'N/A',
                'instaladores': [inst.get_full_name() or inst.username for inst in s.instaladores.all()[:3]],
            })
    
    estadisticas = {
        'total_soportes': total_registros,
        'completados': soportes.filter(estado='COMPLETADO').count(),
        'pendientes': soportes.filter(estado='PENDIENTE').count(),
        'en_proceso': soportes.filter(estado='EN_PROCESO').count(),
    }
    
    return JsonResponse({
        'data': data_list,
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })

@login_required
@user_passes_test(es_admin)
def reporte_inventario_json(request):
    """API para obtener datos del inventario global"""
    
    tipo_reporte = request.GET.get('tipo', 'simple')
    material_id = request.GET.get('material', '')
    busqueda = request.GET.get('busqueda', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    
    inventario = InventarioGlobal.objects.select_related('material', 'actualizado_por')
    
    if material_id:
        inventario = inventario.filter(material_id=material_id)
    if busqueda:
        inventario = inventario.filter(material__nombre__icontains=busqueda)
    
    total_registros = inventario.count()
    paginator = Paginator(inventario, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    if tipo_reporte == 'simple':
        data_list = [{
            'id': item.id,
            'material': item.material.nombre,
            'cantidad': item.cantidad,
            'minimo': item.cantidad_minima,
            'estado': 'Bajo stock' if item.esta_bajo_stock else 'Normal',
        } for item in page_obj]
    else:
        data_list = [{
            'id': item.id,
            'material': item.material.nombre,
            'cantidad': item.cantidad,
            'minimo': item.cantidad_minima,
            'estado': 'Bajo stock' if item.esta_bajo_stock else 'Normal',
            'ultima_actualizacion': item.ultima_actualizacion.strftime('%d/%m/%Y %H:%M'),
            'actualizado_por': item.actualizado_por.get_full_name() or item.actualizado_por.username if item.actualizado_por else 'Sistema',
        } for item in page_obj]
    
    estadisticas = {
        'total_materiales': total_registros,
        'total_unidades': sum(item.cantidad for item in inventario),
        'bajo_stock': sum(1 for item in inventario if item.esta_bajo_stock),
    }
    
    return JsonResponse({
        'data': data_list,
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })


@login_required
@user_passes_test(es_admin)
def reporte_vendedores_json(request):
    """
    API para obtener reporte de vendedores con contratos completados
    por semana (viernes a jueves)
    
    NUEVA LÓGICA DE COMISIONES:
    - Plan 300 Mbps: $8 (normal) / $12 (con cashea)
    - Plan 400 Mbps: $12 (normal) / $15 (con cashea)  
    - Plan 500 Mbps o más: $15 (normal) / $17 (con cashea)
    - BONO: $25 si alcanza 8 o más contratos (SOLO cuentan contratos de 400 Mbps o más)
    """
    
    import pytz
    from decimal import Decimal
    from datetime import datetime, timedelta
    
    # Zona horaria de Venezuela
    VE_TZ = pytz.timezone('America/Caracas')
    
    # Obtener parámetros
    semana_fecha = request.GET.get('semana', '')
    vendedor_id = request.GET.get('vendedor', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    busqueda = request.GET.get('busqueda', '')
    
    # ========== CONVERTIR FECHA DE SEMANA A DATETIME AWARE ==========
    fecha_referencia = None
    if semana_fecha:
        fecha_referencia_aware = convertir_a_datetime_aware(semana_fecha)
        if fecha_referencia_aware:
            fecha_referencia = fecha_referencia_aware.date()
    
    if not fecha_referencia:
        ahora_ve = datetime.now().astimezone(VE_TZ)
        fecha_referencia = ahora_ve.date()
    
    # Calcular semana (viernes a jueves)
    dias_desde_viernes = fecha_referencia.weekday() - 4
    if dias_desde_viernes < 0:
        dias_desde_viernes += 7
    
    viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
    jueves_fin = viernes_inicio + timedelta(days=6)
    
    # ===== CONVERTIR A DATETIME AWARE PARA FILTRAR =====
    fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
    fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
    
    # Contratos completados en la semana
    contratos = ContratoCliente.objects.filter(
        estado='COMPLETADO',
        fecha_completado__gte=fecha_inicio_aware,
        fecha_completado__lte=fecha_fin_aware
    )
    
    # Filtrar por vendedor
    if vendedor_id:
        contratos = contratos.filter(creado_por_id=vendedor_id)
    
    # Obtener todos los vendedores
    todos_vendedores = User.objects.filter(
        groups__name__in=['Vendedor', 'Supervisor', 'Administrador']
    ).distinct().order_by('first_name', 'username')
    
    if vendedor_id:
        todos_vendedores = todos_vendedores.filter(id=vendedor_id)
    
    if busqueda:
        todos_vendedores = todos_vendedores.filter(
            Q(first_name__icontains=busqueda) |
            Q(username__icontains=busqueda) |
            Q(last_name__icontains=busqueda)
        )
    
    # Obtener tasa de cambio
    tasa_obj = TasaCambio.objects.filter(activo=True).first()
    if tasa_obj:
        tasa = float(tasa_obj.tasa)
        tasa_decimal = tasa_obj.tasa
    else:
        tasa = 0
        tasa_decimal = Decimal('0')
    
    # ===== FUNCIÓN PARA CALCULAR COMISIÓN POR PLAN =====
    def calcular_comision_contrato(plan_nombre, cashea):
        """
        Calcula la comisión por contrato según el plan y si tiene cashea
        
        Planes:
        - 300 Mbps: $8 normal / $12 cashea
        - 400 Mbps: $12 normal / $15 cashea
        - 500 Mbps o más: $15 normal / $17 cashea
        """
        # Extraer número del plan
        numeros = re.findall(r'\d+', plan_nombre)
        if not numeros:
            return 0
        
        velocidad = int(numeros[0])
        
        if velocidad == 300:
            return 12 if cashea else 8
        elif velocidad == 400:
            return 15 if cashea else 12
        elif velocidad >= 500:
            return 17 if cashea else 15
        else:
            # Planes menores a 300 Mbps (si existen)
            return 8 if cashea else 5
    
    vendedores_data = []
    
    for vendedor in todos_vendedores:
        contratos_vendedor = contratos.filter(creado_por=vendedor)
        
        # Calcular total de contratos y comisiones
        total_contratos = 0
        total_comision = 0
        contratos_para_bono = 0  # Solo contratos de 400 Mbps o más
        lista_contratos = []
        
        for contrato in contratos_vendedor.order_by('-fecha_completado'):
            plan_nombre = contrato.plan_contratado.nombre
            cashea = contrato.cashea  # True o False
            comision = calcular_comision_contrato(plan_nombre, cashea)
            
            total_contratos += 1
            total_comision += comision
            
            # Verificar si el contrato cuenta para el bono (400 Mbps o más)
            numeros = re.findall(r'\d+', plan_nombre)
            if numeros and int(numeros[0]) >= 400:
                contratos_para_bono += 1
            
            # Convertir fechas a Venezuela
            fecha_completado_ve = contrato.fecha_completado.astimezone(VE_TZ) if contrato.fecha_completado else None
            fecha_completado_str = fecha_completado_ve.strftime('%d/%m/%Y %H:%M') if fecha_completado_ve else 'N/A'
            fecha_creacion_ve = contrato.fecha_creacion.astimezone(VE_TZ) if contrato.fecha_creacion else None
            fecha_creacion_str = fecha_creacion_ve.strftime('%d/%m/%Y') if fecha_creacion_ve else 'N/A'
            
            lista_contratos.append({
                'id': contrato.id,
                'cliente': contrato.nombre_completo,
                'fecha_completado': fecha_completado_str,
                'fecha_creacion': fecha_creacion_str,
                'plan': plan_nombre,
                'customer_id': contrato.customer_id or 'N/A',
                'cashea': 'Sí' if cashea else 'No',
                'comision': f"${comision:.2f}"
            })
        
        # Calcular bono (solo si tiene 8 o más contratos de 400 Mbps o más)
        bono = 25 if contratos_para_bono >= 8 else 0
        total_con_bono = total_comision + bono
        
        # Mostrar información si tiene contratos O si se está filtrando por vendedor específico
        if total_contratos > 0 or vendedor_id:
            # Determinar rango de contratos para el bono
            if contratos_para_bono >= 8:
                rango_bono = f"✅ {contratos_para_bono} contratos (400+ Mbps) - Bono aplicado"
            else:
                rango_bono = f"❌ {contratos_para_bono}/8 contratos (400+ Mbps) - Faltan {8 - contratos_para_bono} para bono"
            
            vendedores_data.append({
                'id': vendedor.id,
                'vendedor': vendedor.get_full_name() or vendedor.username,
                'username': vendedor.username,
                'contratos': total_contratos,
                'contratos_para_bono': contratos_para_bono,
                'comision_total': f"${total_comision:.2f}",
                'bono': f"${bono:.2f}",
                'total_con_bono': f"${total_con_bono:.2f}",
                'total_bs': f"Bs {(total_con_bono * tasa):,.2f}",
                'rango_bono': rango_bono,
                'contratos_detalle': lista_contratos
            })
    
    # Ordenar por total de contratos (de mayor a menor)
    vendedores_data.sort(key=lambda x: x['contratos'], reverse=True)
    
    total_registros = len(vendedores_data)
    paginator = Paginator(vendedores_data, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    # Estadísticas generales
    total_contratos_semana = sum(v['contratos'] for v in vendedores_data)
    total_comisiones_usd = sum(float(v['comision_total'].replace('$', '')) for v in vendedores_data)
    total_bonos_usd = sum(float(v['bono'].replace('$', '')) for v in vendedores_data)
    total_pagar_usd = sum(float(v['total_con_bono'].replace('$', '')) for v in vendedores_data)
    total_pagar_bs = total_pagar_usd * tasa
    
    fecha_tasa = tasa_obj.fecha.strftime('%d/%m/%Y') if tasa_obj else 'No definida'
    tasa_str = f"{float(tasa_decimal):,.2f}" if tasa_decimal else "0.00"
    
    # Calcular vendedores con bono
    vendedores_con_bono = sum(1 for v in vendedores_data if float(v['bono'].replace('$', '')) > 0)
    
    estadisticas = {
        'total_vendedores': len(vendedores_data),
        'vendedores_con_bono': vendedores_con_bono,
        'total_contratos_semana': total_contratos_semana,
        'total_comisiones_usd': f"${total_comisiones_usd:,.2f}",
        'total_bonos_usd': f"${total_bonos_usd:,.2f}",
        'total_pagar_usd': f"${total_pagar_usd:,.2f}",
        'total_pagar_bs': f"Bs {total_pagar_bs:,.2f}",
        'semana_inicio': viernes_inicio.strftime('%d/%m/%Y'),
        'semana_fin': jueves_fin.strftime('%d/%m/%Y'),
        'tasa_cambio': f"1 USD = {tasa_str} Bs",
        'fecha_actualizacion_tasa': fecha_tasa,
        'meta_bono': "8 contratos (400 Mbps o más) para bono de $25"
    }
    
    return JsonResponse({
        'data': list(page_obj),
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
        'semana_inicio': viernes_inicio.strftime('%Y-%m-%d'),
        'semana_fin': jueves_fin.strftime('%Y-%m-%d')
    })


@login_required
@user_passes_test(es_admin)
def semanas_disponibles_api(request):
    """API para obtener las semanas disponibles con contratos completados"""
    
    import pytz
    from datetime import datetime, timedelta
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    contratos = ContratoCliente.objects.filter(estado='COMPLETADO').order_by('fecha_creacion')
    
    semanas = []
    fechas_procesadas = set()
    
    for contrato in contratos:
        if contrato.fecha_completado:
            fecha_utc = contrato.fecha_completado
        else:
            fecha_utc = contrato.fecha_creacion
        
        fecha_ve = fecha_utc.astimezone(VE_TZ)
        fecha = fecha_ve.date()
        
        dias_desde_viernes = fecha.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        clave = viernes_inicio.strftime('%Y-%m-%d')
        
        if clave not in fechas_procesadas:
            fechas_procesadas.add(clave)
            semanas.append({
                'value': viernes_inicio.strftime('%Y-%m-%d'),
                'label': f"{viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"
            })
    
    semanas.sort(key=lambda x: x['value'], reverse=True)
    
    ahora_ve = datetime.now().astimezone(VE_TZ)
    hoy = ahora_ve.date()
    dias_desde_viernes = hoy.weekday() - 4
    if dias_desde_viernes < 0:
        dias_desde_viernes += 7
    
    viernes_actual = hoy - timedelta(days=dias_desde_viernes)
    jueves_actual = viernes_actual + timedelta(days=6)
    semana_actual_clave = viernes_actual.strftime('%Y-%m-%d')
    
    semana_actual = {
        'value': semana_actual_clave,
        'label': f"Semana Actual ({viernes_actual.strftime('%d/%m/%Y')} - {jueves_actual.strftime('%d/%m/%Y')})"
    }
    
    if not any(s['value'] == semana_actual_clave for s in semanas):
        semanas.insert(0, semana_actual)
    
    return JsonResponse({'semanas': semanas})


@staff_member_required
@csrf_exempt
def api_actualizar_tasa(request):
    """API para actualizar la tasa de cambio manualmente"""
    if request.method == 'POST':
        try:
            # Ejecutar el comando
            call_command('actualizar_tasa')
            return JsonResponse({
                'success': True,
                'message': 'Tasa actualizada correctamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)




@login_required
@user_passes_test(es_admin)
def reporte_instaladores_json(request):
    """
    API para obtener reporte de CUADRILLAS para PAGO DE NÓMINA:
    - Instalaciones COMPLETADAS: $15 c/u (usando fecha_instalacion)
    - Soportes COMPLETADOS: según tipo (usando fecha_creacion)
    - Contratos (ventas): $10 c/u - por contratos COMPLETADOS creados por instaladores
    Por semana (viernes a jueves)
    """
    
    import pytz
    from decimal import Decimal
    from collections import defaultdict
    from datetime import datetime, timedelta
    
    # Zona horaria de Venezuela
    VE_TZ = pytz.timezone('America/Caracas')
    
    # Obtener parámetros
    semana_fecha = request.GET.get('semana', '')
    instalador_id = request.GET.get('instalador', '')
    cuadrilla_id = request.GET.get('cuadrilla', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    busqueda = request.GET.get('busqueda', '')
    
    # ========== CONVERTIR FECHA DE SEMANA A DATETIME AWARE ==========
    fecha_referencia = None
    if semana_fecha:
        fecha_referencia_aware = convertir_a_datetime_aware(semana_fecha)
        if fecha_referencia_aware:
            fecha_referencia = fecha_referencia_aware.date()
    
    if not fecha_referencia:
        ahora_ve = datetime.now().astimezone(VE_TZ)
        fecha_referencia = ahora_ve.date()
    
    # Calcular semana (viernes a jueves)
    dias_desde_viernes = fecha_referencia.weekday() - 4
    if dias_desde_viernes < 0:
        dias_desde_viernes += 7
    
    viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
    jueves_fin = viernes_inicio + timedelta(days=6)
    
    # ===== CONVERTIR A DATETIME AWARE PARA FILTRAR =====
    fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
    fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
    
    # Precios
    PRECIO_INSTALACION = 15
    PRECIO_CONTRATO = 10
    PRECIOS_SOPORTES = {
        'SOPORTE': 10,
        'RETIRO': 8,
        'MUDANZA': 15,
        'RECABLEADO': 15,
    }
    
    # Obtener todas las cuadrillas activas
    todas_cuadrillas = Cuadrilla.objects.filter(activo=True)
    
    if cuadrilla_id:
        todas_cuadrillas = todas_cuadrillas.filter(id=cuadrilla_id)
    
    cuadrillas_dict = defaultdict(lambda: {
        'id': None,
        'cuadrilla': '',
        'instalaciones': 0,
        'monto_instalaciones': 0,
        'soportes': 0,
        'monto_soportes': 0,
        'contratos': 0,
        'monto_contratos': 0,
        'instaladores_set': set(),
        'instalaciones_detalle': [],
        'soportes_detalle': [],
        'contratos_detalle': []
    })
    
    for cuadrilla in todas_cuadrillas:
        cuadrillas_dict[cuadrilla.nombre]['id'] = cuadrilla.id
        cuadrillas_dict[cuadrilla.nombre]['cuadrilla'] = cuadrilla.nombre
        
        perfiles = cuadrilla.instaladores.all()
        for perfil in perfiles:
            if perfil.usuario:
                cuadrillas_dict[cuadrilla.nombre]['instaladores_set'].add(perfil.usuario.id)
    
    # ========== 1. INSTALACIONES COMPLETADAS ==========
    instalaciones = Instalacion.objects.filter(
        completada=True,
        fecha_instalacion__gte=fecha_inicio_aware,
        fecha_instalacion__lte=fecha_fin_aware
    ).select_related('asignacion__cuadrilla')
    
    for inst in instalaciones:
        cuadrilla_obj = inst.asignacion.cuadrilla if inst.asignacion else None
        if not cuadrilla_obj:
            continue
        
        nombre_cuadrilla = cuadrilla_obj.nombre
        if nombre_cuadrilla not in cuadrillas_dict:
            continue
        
        instaladores_hist = inst.instaladores.all()
        
        if instalador_id:
            if not instaladores_hist.filter(id=instalador_id).exists():
                continue
        
        if busqueda:
            tiene_coincidencia = False
            for inst_hist in instaladores_hist:
                nombre_completo = inst_hist.get_full_name() or inst_hist.username
                if busqueda.lower() in nombre_completo.lower():
                    tiene_coincidencia = True
                    break
            if not tiene_coincidencia:
                continue
        
        cuadrillas_dict[nombre_cuadrilla]['instalaciones'] += 1
        cuadrillas_dict[nombre_cuadrilla]['monto_instalaciones'] += PRECIO_INSTALACION
        
        for inst_hist in instaladores_hist:
            cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        if len(cuadrillas_dict[nombre_cuadrilla]['instalaciones_detalle']) < 5:
            cliente_nombre = inst.nombre_cliente if hasattr(inst, 'nombre_cliente') else 'N/A'
            customer_id = inst.customer_id if hasattr(inst, 'customer_id') else 'N/A'
            nombres_inst = [i.get_full_name() or i.username for i in instaladores_hist[:3]]
            
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if inst.fecha_instalacion:
                fecha_ve = inst.fecha_instalacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            else:
                fecha_str = 'N/A'
            
            cuadrillas_dict[nombre_cuadrilla]['instalaciones_detalle'].append({
                'cliente': cliente_nombre,
                'customer_id': customer_id,
                'fecha': fecha_str,
                'instaladores': ', '.join(nombres_inst)
            })
    
    # ========== 2. SOPORTES COMPLETADOS ==========
    soportes = Soporte.objects.filter(
        estado='COMPLETADO',
        fecha_creacion__gte=fecha_inicio_aware,
        fecha_creacion__lte=fecha_fin_aware
    ).select_related('cuadrilla')
    
    for sop in soportes:
        if not sop.cuadrilla:
            continue
        
        nombre_cuadrilla = sop.cuadrilla.nombre
        if nombre_cuadrilla not in cuadrillas_dict:
            continue
        
        instaladores_hist = sop.instaladores.all()
        
        if instalador_id:
            if not instaladores_hist.filter(id=instalador_id).exists():
                continue
        
        if busqueda:
            tiene_coincidencia = False
            for inst_hist in instaladores_hist:
                nombre_completo = inst_hist.get_full_name() or inst_hist.username
                if busqueda.lower() in nombre_completo.lower():
                    tiene_coincidencia = True
                    break
            if not tiene_coincidencia:
                continue
        
        try:
            tipo = sop.asignacion.ticket.tipo_soporte if sop.asignacion and sop.asignacion.ticket else 'SOPORTE'
            precio = PRECIOS_SOPORTES.get(tipo, 10)
        except:
            precio = 10
            tipo = 'SOPORTE'
        
        cuadrillas_dict[nombre_cuadrilla]['soportes'] += 1
        cuadrillas_dict[nombre_cuadrilla]['monto_soportes'] += precio
        
        for inst_hist in instaladores_hist:
            cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        if len(cuadrillas_dict[nombre_cuadrilla]['soportes_detalle']) < 5:
            cliente_nombre = 'N/A'
            try:
                if sop.asignacion and sop.asignacion.ticket:
                    cliente_nombre = sop.asignacion.ticket.nombre_completo
            except:
                pass
            nombres_inst = [i.get_full_name() or i.username for i in instaladores_hist[:3]]
            
            # ===== CONVERTIR FECHA A VENEZUELA =====
            if sop.fecha_creacion:
                fecha_ve = sop.fecha_creacion.astimezone(VE_TZ)
                fecha_str = fecha_ve.strftime('%d/%m/%Y')
            else:
                fecha_str = 'N/A'
            
            cuadrillas_dict[nombre_cuadrilla]['soportes_detalle'].append({
                'ticket_padre': sop.asignacion.ticket.ticket_padre if sop.asignacion and sop.asignacion.ticket else 'N/A',
                'cliente': cliente_nombre,
                'tipo': tipo,
                'precio': precio,
                'fecha': fecha_str,
                'instaladores': ', '.join(nombres_inst)
            })
    
    # ========== 3. CONTRATOS COMPLETADOS ==========
    instaladores_users = User.objects.filter(groups__name='Instalador')
    
    if instalador_id:
        instaladores_users = instaladores_users.filter(id=instalador_id)
    
    for instalador in instaladores_users:
        perfil = PerfilUsuario.objects.filter(usuario=instalador).first()
        if not perfil:
            continue
        
        cuadrillas_del_instalador = perfil.cuadrillas.all()
        if not cuadrillas_del_instalador.exists():
            continue
        
        contratos_instalador = ContratoCliente.objects.filter(
            estado='COMPLETADO',
            creado_por=instalador,
            fecha_completado__gte=fecha_inicio_aware,
            fecha_completado__lte=fecha_fin_aware
        )
        
        if busqueda:
            contratos_instalador = contratos_instalador.filter(
                Q(cliente_potencial__nombre__icontains=busqueda) |
                Q(cliente_potencial__apellido__icontains=busqueda) |
                Q(cliente_potencial__cedula__icontains=busqueda) |
                Q(customer_id__icontains=busqueda)
            )
        
        for contrato in contratos_instalador:
            for cuadrilla_inst in cuadrillas_del_instalador:
                nombre_cuadrilla = cuadrilla_inst.nombre
                if nombre_cuadrilla not in cuadrillas_dict:
                    continue
                
                if cuadrilla_id and int(cuadrilla_id) != cuadrillas_dict[nombre_cuadrilla]['id']:
                    continue
                
                cuadrillas_dict[nombre_cuadrilla]['contratos'] += 1
                cuadrillas_dict[nombre_cuadrilla]['monto_contratos'] += PRECIO_CONTRATO
                cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(instalador.id)
                
                if len(cuadrillas_dict[nombre_cuadrilla]['contratos_detalle']) < 5:
                    # ===== CONVERTIR FECHA A VENEZUELA =====
                    if contrato.fecha_completado:
                        fecha_ve = contrato.fecha_completado.astimezone(VE_TZ)
                        fecha_str = fecha_ve.strftime('%d/%m/%Y')
                    else:
                        fecha_str = 'N/A'
                    
                    cuadrillas_dict[nombre_cuadrilla]['contratos_detalle'].append({
                        'cliente': contrato.nombre_completo,
                        'plan': contrato.plan_contratado.nombre,
                        'customer_id': contrato.customer_id or 'N/A',
                        'fecha_completado': fecha_str
                    })
    
    cuadrillas_data = []
    for nombre, data in cuadrillas_dict.items():
        total_usd = data['monto_instalaciones'] + data['monto_soportes'] + data['monto_contratos']
        
        nombres_instaladores = []
        for inst_id in data['instaladores_set']:
            try:
                inst = User.objects.get(id=inst_id)
                nombres_instaladores.append(inst.get_full_name() or inst.username)
            except:
                pass
        
        if (data['instalaciones'] > 0 or data['soportes'] > 0 or data['contratos'] > 0 or cuadrilla_id):
            cuadrillas_data.append({
                'id': data['id'],
                'cuadrilla': data['cuadrilla'],
                'instalaciones': data['instalaciones'],
                'monto_instalaciones': f"${data['monto_instalaciones']}",
                'soportes': data['soportes'],
                'monto_soportes': f"${data['monto_soportes']}",
                'contratos': data['contratos'],
                'monto_contratos': f"${data['monto_contratos']}",
                'total_usd': f"${total_usd}",
                'instaladores_list': nombres_instaladores,
                'instalaciones_detalle': data['instalaciones_detalle'],
                'soportes_detalle': data['soportes_detalle'],
                'contratos_detalle': data['contratos_detalle'],
            })
    
    cuadrillas_data.sort(key=lambda x: x['instalaciones'] + x['soportes'] + x['contratos'], reverse=True)
    
    tasa_obj = TasaCambio.objects.filter(activo=True).first()
    if tasa_obj:
        tasa = float(tasa_obj.tasa)
        tasa_decimal = tasa_obj.tasa
    else:
        tasa = 0
        tasa_decimal = Decimal('0')
    
    for item in cuadrillas_data:
        total_usd_num = float(item['total_usd'].replace('$', ''))
        item['total_bs'] = f"Bs {total_usd_num * tasa:,.2f}"
    
    total_registros = len(cuadrillas_data)
    paginator = Paginator(cuadrillas_data, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    total_instalaciones_semana = sum(v['instalaciones'] for v in cuadrillas_data)
    total_soportes_semana = sum(v['soportes'] for v in cuadrillas_data)
    total_contratos_semana = sum(v['contratos'] for v in cuadrillas_data)
    total_pagar_usd = sum(float(v['total_usd'].replace('$', '')) for v in cuadrillas_data)
    total_pagar_bs = total_pagar_usd * tasa
    
    fecha_tasa = tasa_obj.fecha.strftime('%d/%m/%Y') if tasa_obj else 'No definida'
    tasa_str = f"{float(tasa_decimal):,.2f}" if tasa_decimal else "0.00"
    
    estadisticas = {
        'total_cuadrillas': len(cuadrillas_data),
        'total_instalaciones_semana': total_instalaciones_semana,
        'total_soportes_semana': total_soportes_semana,
        'total_contratos_semana': total_contratos_semana,
        'total_pagar_usd': f"${total_pagar_usd:,.2f}",
        'total_pagar_bs': f"Bs {total_pagar_bs:,.2f}",
        'semana_inicio': viernes_inicio.strftime('%d/%m/%Y'),
        'semana_fin': jueves_fin.strftime('%d/%m/%Y'),
        'tasa_cambio': f"1 USD = {tasa_str} Bs",
        'fecha_actualizacion_tasa': fecha_tasa,
        'precio_instalacion': f"${PRECIO_INSTALACION}",
        'precio_contrato': f"${PRECIO_CONTRATO}",
    }
    
    return JsonResponse({
        'data': list(page_obj),
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
        'semana_inicio': viernes_inicio.strftime('%Y-%m-%d'),
        'semana_fin': jueves_fin.strftime('%Y-%m-%d')
    })


@login_required
@user_passes_test(es_admin)
def reporte_simple_vendedores_json(request):
    """
    API para reporte simple de vendedores - Cantidad de contratos por vendedor
    Filtros por fechas y vendedor específico
    """
    
    import pytz
    from decimal import Decimal
    from datetime import datetime, timedelta
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    # Obtener parámetros
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    vendedor_id = request.GET.get('vendedor', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    busqueda = request.GET.get('busqueda', '')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    # Contratos completados en el período
    contratos = ContratoCliente.objects.filter(estado='COMPLETADO')
    
    if fecha_desde_aware and fecha_hasta_aware:
        contratos = contratos.filter(fecha_completado__gte=fecha_desde_aware, fecha_completado__lte=fecha_hasta_aware)
    elif fecha_desde_aware:
        contratos = contratos.filter(fecha_completado__gte=fecha_desde_aware)
    elif fecha_hasta_aware:
        contratos = contratos.filter(fecha_completado__lte=fecha_hasta_aware)
    
    if vendedor_id:
        contratos = contratos.filter(creado_por_id=vendedor_id)
    
    # Obtener todos los vendedores
    vendedores_list = User.objects.filter(
        groups__name__in=['Vendedor', 'Supervisor', 'Administrador']
    ).distinct().order_by('first_name', 'username')
    
    if vendedor_id:
        vendedores_list = vendedores_list.filter(id=vendedor_id)
    
    if busqueda:
        vendedores_list = vendedores_list.filter(
            Q(first_name__icontains=busqueda) |
            Q(username__icontains=busqueda) |
            Q(last_name__icontains=busqueda)
        )
    
    data_vendedores = []
    
    for vendedor in vendedores_list:
        contratos_vendedor = contratos.filter(creado_por=vendedor)
        total_contratos = contratos_vendedor.count()
        
        if total_contratos > 0 or vendedor_id:
            # Obtener detalle de contratos
            contratos_detalle = []
            for contrato in contratos_vendedor.order_by('-fecha_completado')[:10]:
                fecha_ve = contrato.fecha_completado.astimezone(VE_TZ) if contrato.fecha_completado else None
                fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
                
                contratos_detalle.append({
                    'id': contrato.id,
                    'cliente': contrato.nombre_completo,
                    'fecha': fecha_str,
                    'plan': contrato.plan_contratado.nombre,
                    'customer_id': contrato.customer_id or 'N/A'
                })
            
            data_vendedores.append({
                'id': vendedor.id,
                'vendedor': vendedor.get_full_name() or vendedor.username,
                'username': vendedor.username,
                'contratos': total_contratos,
                'contratos_detalle': contratos_detalle
            })
    
    data_vendedores.sort(key=lambda x: x['contratos'], reverse=True)
    
    total_registros = len(data_vendedores)
    paginator = Paginator(data_vendedores, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    total_contratos_general = sum(v['contratos'] for v in data_vendedores)
    
    estadisticas = {
        'total_vendedores': len(data_vendedores),
        'total_contratos': total_contratos_general,
        'promedio_contratos': round(total_contratos_general / len(data_vendedores), 2) if data_vendedores else 0,
        'fecha_desde': fecha_desde_aware.strftime('%d/%m/%Y') if fecha_desde_aware else 'Todo',
        'fecha_hasta': fecha_hasta_aware.strftime('%d/%m/%Y') if fecha_hasta_aware else 'Actual'
    }
    
    return JsonResponse({
        'data': list(page_obj),
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })


@login_required
@user_passes_test(es_admin)
def reporte_global_json(request):
    """
    API para reporte global - Ventas directas, Contratos y Soportes
    Filtros: fechas, vendedor, instalador, cuadrilla
    """
    
    import pytz
    from decimal import Decimal
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    # Obtener parámetros
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    vendedor_id = request.GET.get('vendedor', '')
    instalador_id = request.GET.get('instalador', '')
    cuadrilla_id = request.GET.get('cuadrilla', '')
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 15))
    busqueda = request.GET.get('busqueda', '')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    # 1. VENTAS DIRECTAS
    ventas_directas = VentaDirecta.objects.filter(estado='COMPLETADO')
    if fecha_desde_aware and fecha_hasta_aware:
        ventas_directas = ventas_directas.filter(fecha_creacion__gte=fecha_desde_aware, fecha_creacion__lte=fecha_hasta_aware)
    elif fecha_desde_aware:
        ventas_directas = ventas_directas.filter(fecha_creacion__gte=fecha_desde_aware)
    elif fecha_hasta_aware:
        ventas_directas = ventas_directas.filter(fecha_creacion__lte=fecha_hasta_aware)
    
    if vendedor_id:
        ventas_directas = ventas_directas.filter(creado_por_id=vendedor_id)
    
    # 2. CONTRATOS (Ventas de vendedores)
    contratos = ContratoCliente.objects.filter(estado='COMPLETADO')
    if fecha_desde_aware and fecha_hasta_aware:
        contratos = contratos.filter(fecha_completado__gte=fecha_desde_aware, fecha_completado__lte=fecha_hasta_aware)
    elif fecha_desde_aware:
        contratos = contratos.filter(fecha_completado__gte=fecha_desde_aware)
    elif fecha_hasta_aware:
        contratos = contratos.filter(fecha_completado__lte=fecha_hasta_aware)
    
    if vendedor_id:
        contratos = contratos.filter(creado_por_id=vendedor_id)
    
    # 3. INSTALACIONES
    instalaciones = Instalacion.objects.filter(completada=True)
    if fecha_desde_aware and fecha_hasta_aware:
        instalaciones = instalaciones.filter(fecha_instalacion__gte=fecha_desde_aware, fecha_instalacion__lte=fecha_hasta_aware)
    elif fecha_desde_aware:
        instalaciones = instalaciones.filter(fecha_instalacion__gte=fecha_desde_aware)
    elif fecha_hasta_aware:
        instalaciones = instalaciones.filter(fecha_instalacion__lte=fecha_hasta_aware)
    
    if cuadrilla_id:
        instalaciones = instalaciones.filter(asignacion__cuadrilla_id=cuadrilla_id)
    
    if instalador_id:
        instalaciones = instalaciones.filter(instaladores__id=instalador_id)
    
    # 4. SOPORTES (tickets completados)
    soportes = Soporte.objects.filter(estado='COMPLETADO')
    if fecha_desde_aware and fecha_hasta_aware:
        soportes = soportes.filter(fecha_creacion__gte=fecha_desde_aware, fecha_creacion__lte=fecha_hasta_aware)
    elif fecha_desde_aware:
        soportes = soportes.filter(fecha_creacion__gte=fecha_desde_aware)
    elif fecha_hasta_aware:
        soportes = soportes.filter(fecha_creacion__lte=fecha_hasta_aware)
    
    if cuadrilla_id:
        soportes = soportes.filter(cuadrilla_id=cuadrilla_id)
    
    if instalador_id:
        soportes = soportes.filter(instaladores__id=instalador_id)
    
    # Buscar por cliente/ticket
    if busqueda:
        contratos = contratos.filter(
            Q(cliente_potencial__nombre__icontains=busqueda) |
            Q(cliente_potencial__apellido__icontains=busqueda) |
            Q(cliente_potencial__cedula__icontains=busqueda) |
            Q(customer_id__icontains=busqueda)
        )
        ventas_directas = ventas_directas.filter(
            Q(nombre__icontains=busqueda) |
            Q(apellido__icontains=busqueda) |
            Q(cedula__icontains=busqueda) |
            Q(customer_id__icontains=busqueda)
        )
        instalaciones = instalaciones.filter(
            Q(nombre_cliente__icontains=busqueda) |
            Q(cedula_cliente__icontains=busqueda) |
            Q(customer_id__icontains=busqueda)
        )
        soportes = soportes.filter(
            Q(asignacion__ticket__nombre__icontains=busqueda) |
            Q(asignacion__ticket__apellido__icontains=busqueda) |
            Q(asignacion__ticket__cedula__icontains=busqueda) |
            Q(asignacion__ticket__customer_id__icontains=busqueda)
        )
    
    # ========== ESTADÍSTICAS POR PLAN (CONTRATOS) ==========
    planes_stats = defaultdict(int)
    for c in contratos:
        plan_nombre = c.plan_contratado.nombre
        numeros = re.findall(r'\d+', plan_nombre)
        if numeros:
            velocidad = int(numeros[0])
            if velocidad <= 300:
                planes_stats['300 Mbps'] += 1
            elif velocidad == 400:
                planes_stats['400 Mbps'] += 1
            elif velocidad >= 500:
                planes_stats['500+ Mbps'] += 1
            else:
                planes_stats[plan_nombre] += 1
        else:
            planes_stats[plan_nombre] += 1
    
    # ========== ESTADÍSTICAS POR TIPO DE SOPORTE ==========
    soportes_stats = defaultdict(int)
    for s in soportes:
        try:
            tipo = s.asignacion.ticket.tipo_soporte if s.asignacion and s.asignacion.ticket else 'SOPORTE'
            if tipo == 'MUDANZA':
                soportes_stats['Mudanza'] += 1
            elif tipo == 'RETIRO':
                soportes_stats['Retiro'] += 1
            elif tipo == 'RECABLEADO':
                soportes_stats['Recableado'] += 1
            else:
                soportes_stats['Soporte Técnico'] += 1
        except:
            soportes_stats['Soporte Técnico'] += 1
    
    # ========== ESTADÍSTICAS DE MATERIALES (INSTALACIONES) ==========
    materiales_stats = defaultdict(int)
    for i in instalaciones:
        if i.conectores:
            materiales_stats['Conectores'] += i.conectores
        if i.rosetas:
            materiales_stats['Rosetas'] += i.rosetas
        if i.patch_cord:
            materiales_stats['Patch Cord'] += i.patch_cord
        if i.tensores:
            materiales_stats['Tensores'] += i.tensores
        if i.tirros:
            materiales_stats['Tirros'] += i.tirros
        if i.metros_utilizados:
            materiales_stats['Metros de Fibra'] += i.metros_utilizados
    
    # ========== DATOS PARA LA TABLA PRINCIPAL ==========
    registros = []
    
    # Agregar ventas directas
    for vd in ventas_directas.order_by('-fecha_creacion'):
        fecha_ve = vd.fecha_creacion.astimezone(VE_TZ) if vd.fecha_creacion else None
        fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
        
        registros.append({
            'id': vd.id,
            'tipo': 'venta_directa',
            'tipo_display': '🛒 Venta Directa',
            'cliente': vd.nombre_completo,
            'customer_id': vd.customer_id or 'N/A',
            'referencia': vd.nro_orden,
            'fecha': fecha_str,
            'responsable': vd.creado_por.get_full_name() or vd.creado_por.username if vd.creado_por else 'N/A',
            'detalle': vd.plan.nombre,
            'estado': 'Completado'
        })
    
    # Agregar contratos
    for c in contratos.order_by('-fecha_completado'):
        fecha_ve = c.fecha_completado.astimezone(VE_TZ) if c.fecha_completado else None
        fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
        
        registros.append({
            'id': c.id,
            'tipo': 'contrato',
            'tipo_display': '📝 Contrato',
            'cliente': c.nombre_completo,
            'customer_id': c.customer_id or 'N/A',
            'referencia': c.ods or 'N/A',
            'fecha': fecha_str,
            'responsable': c.creado_por.get_full_name() or c.creado_por.username if c.creado_por else 'N/A',
            'detalle': c.plan_contratado.nombre,
            'estado': 'Completado'
        })
    
    # Agregar instalaciones
    for i in instalaciones.order_by('-fecha_instalacion'):
        fecha_ve = i.fecha_instalacion.astimezone(VE_TZ) if i.fecha_instalacion else None
        fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
        
        registros.append({
            'id': i.id,
            'tipo': 'instalacion',
            'tipo_display': '🔧 Instalación',
            'cliente': i.nombre_cliente,
            'customer_id': i.customer_id,
            'referencia': i.orden_servicio,
            'fecha': fecha_str,
            'responsable': i.asignacion.cuadrilla.nombre if i.asignacion and i.asignacion.cuadrilla else 'N/A',
            'detalle': i.plan,
            'estado': 'Completada'
        })
    
    # Agregar soportes
    for s in soportes.order_by('-fecha_creacion'):
        fecha_ve = s.fecha_creacion.astimezone(VE_TZ) if s.fecha_creacion else None
        fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
        
        try:
            ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else 'N/A'
            cliente = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else 'N/A'
            customer_id = s.asignacion.ticket.customer_id if s.asignacion and s.asignacion.ticket else 'N/A'
            tipo_soporte = s.asignacion.ticket.get_tipo_soporte_display() if s.asignacion and s.asignacion.ticket else 'N/A'
        except:
            ticket_padre = 'N/A'
            cliente = 'N/A'
            customer_id = 'N/A'
            tipo_soporte = 'N/A'
        
        registros.append({
            'id': s.id,
            'tipo': 'soporte',
            'tipo_display': f'🔧 Soporte ({tipo_soporte})',
            'cliente': cliente,
            'customer_id': customer_id,
            'referencia': ticket_padre,
            'fecha': fecha_str,
            'responsable': s.cuadrilla.nombre if s.cuadrilla else 'N/A',
            'detalle': s.falla_encontrada[:50] + '...' if s.falla_encontrada and len(s.falla_encontrada) > 50 else (s.falla_encontrada or 'N/A'),
            'estado': s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado
        })
    
    # Ordenar por fecha (más recientes primero)
    registros.sort(key=lambda x: x['fecha'], reverse=True)
    
    total_registros = len(registros)
    paginator = Paginator(registros, per_page)
    
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    # Estadísticas generales
    estadisticas = {
        'total_registros': total_registros,
        'total_ventas_directas': ventas_directas.count(),
        'total_contratos': contratos.count(),
        'total_instalaciones': instalaciones.count(),
        'total_soportes': soportes.count(),
        'planes_stats': dict(planes_stats),
        'soportes_stats': dict(soportes_stats),
        'materiales_stats': dict(materiales_stats),
        'fecha_desde': fecha_desde_aware.strftime('%d/%m/%Y') if fecha_desde_aware else 'Todo',
        'fecha_hasta': fecha_hasta_aware.strftime('%d/%m/%Y') if fecha_hasta_aware else 'Actual'
    }
    
    return JsonResponse({
        'data': list(page_obj),
        'estadisticas': estadisticas,
        'total_registros': total_registros,
        'total_paginas': paginator.num_pages,
        'pagina_actual': page_obj.number,
        'por_pagina': per_page,
    })


@login_required
@user_passes_test(es_admin)
def reporte_global_detalle_json(request):
    """
    API para obtener el resumen estadístico de un tipo específico del reporte global
    Tipos: ventas_directas, contratos, instalaciones, soportes
    """
    
    import pytz
    import re
    from datetime import datetime, timedelta
    from collections import defaultdict
    from django.db.models import Q, Sum
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    # Obtener parámetros
    tipo = request.GET.get('tipo', '')
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    vendedor_id = request.GET.get('vendedor', '')
    instalador_id = request.GET.get('instalador', '')
    cuadrilla_id = request.GET.get('cuadrilla', '')
    
    # ========== CONVERTIR FECHAS A DATETIME AWARE ==========
    fecha_desde_aware = convertir_a_datetime_aware(fecha_desde_raw)
    fecha_hasta_aware = convertir_a_datetime_aware(fecha_hasta_raw)
    
    if fecha_hasta_aware:
        fecha_hasta_aware = (fecha_hasta_aware + timedelta(days=1)) - timedelta(microseconds=1)
    
    estadisticas = {}
    
    if tipo == 'ventas_directas':
        # Ventas Directas
        queryset = VentaDirecta.objects.filter(estado='COMPLETADO')
        
        if fecha_desde_aware and fecha_hasta_aware:
            queryset = queryset.filter(fecha_creacion__gte=fecha_desde_aware, fecha_creacion__lte=fecha_hasta_aware)
        elif fecha_desde_aware:
            queryset = queryset.filter(fecha_creacion__gte=fecha_desde_aware)
        elif fecha_hasta_aware:
            queryset = queryset.filter(fecha_creacion__lte=fecha_hasta_aware)
        
        if vendedor_id:
            queryset = queryset.filter(creado_por_id=vendedor_id)
        
        # Estadísticas por plan
        planes_stats = defaultdict(int)
        for vd in queryset:
            plan_nombre = vd.plan.nombre
            numeros = re.findall(r'\d+', plan_nombre)
            if numeros:
                velocidad = int(numeros[0])
                if velocidad <= 300:
                    plan_key = '300 Mbps'
                elif velocidad == 400:
                    plan_key = '400 Mbps'
                elif velocidad >= 500:
                    plan_key = '500+ Mbps'
                else:
                    plan_key = plan_nombre
            else:
                plan_key = plan_nombre
            planes_stats[plan_key] += 1
        
        estadisticas = {
            'titulo': '🛒 Resumen de Ventas Directas',
            'total': queryset.count(),
            'detalles': [{'nombre': k, 'cantidad': v, 'porcentaje': round((v/queryset.count())*100, 1) if queryset.count() > 0 else 0} for k, v in planes_stats.items()]
        }
    
    elif tipo == 'contratos':
        # Contratos
        queryset = ContratoCliente.objects.filter(estado='COMPLETADO')
        
        if fecha_desde_aware and fecha_hasta_aware:
            queryset = queryset.filter(fecha_completado__gte=fecha_desde_aware, fecha_completado__lte=fecha_hasta_aware)
        elif fecha_desde_aware:
            queryset = queryset.filter(fecha_completado__gte=fecha_desde_aware)
        elif fecha_hasta_aware:
            queryset = queryset.filter(fecha_completado__lte=fecha_hasta_aware)
        
        if vendedor_id:
            queryset = queryset.filter(creado_por_id=vendedor_id)
        
        # Estadísticas por plan
        planes_stats = defaultdict(int)
        for c in queryset:
            plan_nombre = c.plan_contratado.nombre
            numeros = re.findall(r'\d+', plan_nombre)
            if numeros:
                velocidad = int(numeros[0])
                if velocidad <= 300:
                    plan_key = '300 Mbps'
                elif velocidad == 400:
                    plan_key = '400 Mbps'
                elif velocidad >= 500:
                    plan_key = '500+ Mbps'
                else:
                    plan_key = plan_nombre
            else:
                plan_key = plan_nombre
            planes_stats[plan_key] += 1
        
        estadisticas = {
            'titulo': '📝 Resumen de Contratos',
            'total': queryset.count(),
            'detalles': [{'nombre': k, 'cantidad': v, 'porcentaje': round((v/queryset.count())*100, 1) if queryset.count() > 0 else 0} for k, v in planes_stats.items()]
        }
    
    elif tipo == 'instalaciones':
        # Instalaciones
        queryset = Instalacion.objects.filter(completada=True)
        
        if fecha_desde_aware and fecha_hasta_aware:
            queryset = queryset.filter(fecha_instalacion__gte=fecha_desde_aware, fecha_instalacion__lte=fecha_hasta_aware)
        elif fecha_desde_aware:
            queryset = queryset.filter(fecha_instalacion__gte=fecha_desde_aware)
        elif fecha_hasta_aware:
            queryset = queryset.filter(fecha_instalacion__lte=fecha_hasta_aware)
        
        if cuadrilla_id:
            queryset = queryset.filter(asignacion__cuadrilla_id=cuadrilla_id)
        
        if instalador_id:
            queryset = queryset.filter(instaladores__id=instalador_id)
        
        # Estadísticas de materiales
        materiales_stats = defaultdict(int)
        for inst in queryset:
            if inst.conectores:
                materiales_stats['Conectores'] += inst.conectores
            if inst.rosetas:
                materiales_stats['Rosetas'] += inst.rosetas
            if inst.patch_cord:
                materiales_stats['Patch Cord'] += inst.patch_cord
            if inst.tensores:
                materiales_stats['Tensores'] += inst.tensores
            if inst.tirros:
                materiales_stats['Tirros'] += inst.tirros
            if inst.metros_utilizados:
                materiales_stats['Metros de Fibra'] += inst.metros_utilizados
        
        # Estadísticas por plan en instalaciones
        planes_stats = defaultdict(int)
        for inst in queryset:
            plan_nombre = inst.plan
            numeros = re.findall(r'\d+', plan_nombre)
            if numeros:
                velocidad = int(numeros[0])
                if velocidad <= 300:
                    plan_key = '300 Mbps'
                elif velocidad == 400:
                    plan_key = '400 Mbps'
                elif velocidad >= 500:
                    plan_key = '500+ Mbps'
                else:
                    plan_key = plan_nombre
            else:
                plan_key = plan_nombre
            planes_stats[plan_key] += 1
        
        estadisticas = {
            'titulo': '🔧 Resumen de Instalaciones',
            'total': queryset.count(),
            'detalles_planes': [{'nombre': k, 'cantidad': v, 'porcentaje': round((v/queryset.count())*100, 1) if queryset.count() > 0 else 0} for k, v in planes_stats.items()],
            'detalles_materiales': [{'nombre': k, 'cantidad': v} for k, v in materiales_stats.items()]
        }
    
    elif tipo == 'soportes':
        # Soportes
        queryset = Soporte.objects.filter(estado='COMPLETADO')
        
        if fecha_desde_aware and fecha_hasta_aware:
            queryset = queryset.filter(fecha_creacion__gte=fecha_desde_aware, fecha_creacion__lte=fecha_hasta_aware)
        elif fecha_desde_aware:
            queryset = queryset.filter(fecha_creacion__gte=fecha_desde_aware)
        elif fecha_hasta_aware:
            queryset = queryset.filter(fecha_creacion__lte=fecha_hasta_aware)
        
        if cuadrilla_id:
            queryset = queryset.filter(cuadrilla_id=cuadrilla_id)
        
        if instalador_id:
            queryset = queryset.filter(instaladores__id=instalador_id)
        
        # Estadísticas por tipo de soporte
        soportes_stats = defaultdict(int)
        for s in queryset:
            try:
                tipo_soporte = s.asignacion.ticket.get_tipo_soporte_display() if s.asignacion and s.asignacion.ticket else 'Soporte Técnico'
            except:
                tipo_soporte = 'Soporte Técnico'
            soportes_stats[tipo_soporte] += 1
        
        estadisticas = {
            'titulo': '🛠️ Resumen de Soportes',
            'total': queryset.count(),
            'detalles': [{'nombre': k, 'cantidad': v, 'porcentaje': round((v/queryset.count())*100, 1) if queryset.count() > 0 else 0} for k, v in soportes_stats.items()]
        }
    
    return JsonResponse({
        'estadisticas': estadisticas
    })    

@login_required
@user_passes_test(es_admin)
def semanas_disponibles_instaladores_api(request):
    """API para obtener las semanas disponibles con actividad de instaladores"""
    
    import pytz
    from datetime import datetime, timedelta
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    fechas = []
    
    # Instalaciones - convertir a zona Venezuela
    instalaciones = Instalacion.objects.filter(completada=True).exclude(fecha_instalacion__isnull=True)
    for inst in instalaciones:
        if inst.fecha_instalacion:
            fecha_ve = inst.fecha_instalacion.astimezone(VE_TZ)
            fechas.append(fecha_ve.date())
    
    # Soportes - convertir a zona Venezuela
    soportes = Soporte.objects.filter(estado='COMPLETADO').exclude(fecha_hora_servicio__isnull=True)
    for sop in soportes:
        if sop.fecha_hora_servicio:
            fecha_ve = sop.fecha_hora_servicio.astimezone(VE_TZ)
            fechas.append(fecha_ve.date())
    
    # Contratos completados - convertir a zona Venezuela
    contratos = ContratoCliente.objects.filter(estado='COMPLETADO').exclude(fecha_completado__isnull=True)
    for con in contratos:
        if con.fecha_completado:
            fecha_ve = con.fecha_completado.astimezone(VE_TZ)
            fechas.append(fecha_ve.date())
    
    semanas = []
    fechas_procesadas = set()
    
    for fecha in fechas:
        dias_desde_viernes = fecha.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        clave = viernes_inicio.strftime('%Y-%m-%d')
        
        if clave not in fechas_procesadas:
            fechas_procesadas.add(clave)
            semanas.append({
                'value': viernes_inicio.strftime('%Y-%m-%d'),
                'label': f"{viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"
            })
    
    semanas.sort(key=lambda x: x['value'], reverse=True)
    
    ahora_ve = datetime.now().astimezone(VE_TZ)
    hoy = ahora_ve.date()
    dias_desde_viernes = hoy.weekday() - 4
    if dias_desde_viernes < 0:
        dias_desde_viernes += 7
    
    viernes_actual = hoy - timedelta(days=dias_desde_viernes)
    jueves_actual = viernes_actual + timedelta(days=6)
    semana_actual_clave = viernes_actual.strftime('%Y-%m-%d')
    
    semana_actual = {
        'value': semana_actual_clave,
        'label': f"Semana Actual ({viernes_actual.strftime('%d/%m/%Y')} - {jueves_actual.strftime('%d/%m/%Y')})"
    }
    
    if not any(s['value'] == semana_actual_clave for s in semanas):
        semanas.insert(0, semana_actual)
    
    return JsonResponse({'semanas': semanas})   

def convertir_fecha_date(fecha_str):
    """Convierte string a objeto date (sin zona horaria)"""
    if not fecha_str:
        return None
    try:
        return datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        try:
            return datetime.strptime(fecha_str, '%d/%m/%Y').date()
        except ValueError:
            return None

@login_required
@user_passes_test(es_admin)
def exportar_reporte(request):
    """Exportar datos a Excel o PDF"""
    
    formato = request.GET.get('formato', 'excel')
    tipo = request.GET.get('tipo', 'ventas')
    reporte_tipo = request.GET.get('reporte_tipo', 'simple')
    
    fecha_desde_raw = request.GET.get('fecha_desde', '')
    fecha_hasta_raw = request.GET.get('fecha_hasta', '')
    busqueda = request.GET.get('busqueda', '')
    
    vendedor = request.GET.get('vendedor', '')
    plan = request.GET.get('plan', '')
    cuadrilla = request.GET.get('cuadrilla', '')
    estado = request.GET.get('estado', '')  # NUEVO - Filtro de estado para ventas
    tipo_soporte = request.GET.get('tipo_soporte', '')
    estado_soporte = request.GET.get('estado_soporte', '')
    material = request.GET.get('material', '')
    semana_raw = request.GET.get('semana', '')
    instalador = request.GET.get('instalador', '')
    
    # ========== CONVERTIR FECHAS ==========
    fecha_desde_obj = convertir_fecha_date(fecha_desde_raw)
    fecha_hasta_obj = convertir_fecha_date(fecha_hasta_raw)
    
    # Convertir semana si existe
    semana_obj = None
    if semana_raw:
        semana_obj = convertir_fecha_date(semana_raw)
    
    # Pasar a las funciones de exportación como objetos date
    if formato == 'excel':
        return exportar_excel(request, tipo, reporte_tipo, fecha_desde_obj, fecha_hasta_obj, 
                              busqueda, vendedor, plan, cuadrilla, estado, 
                              tipo_soporte, estado_soporte, material, semana_obj, instalador)
    else:
        return exportar_pdf(request, tipo, reporte_tipo, fecha_desde_obj, fecha_hasta_obj,
                            busqueda, vendedor, plan, cuadrilla, estado,
                            tipo_soporte, estado_soporte, material, semana_obj, instalador)


def exportar_excel(request, tipo, reporte_tipo, fecha_desde_obj, fecha_hasta_obj,
                   busqueda, vendedor, plan, cuadrilla, estado,
                   tipo_soporte, estado_soporte, material, semana_obj, instalador):
    """Exportar datos a Excel con filtros usando datetime aware"""
    
    import pytz
    from datetime import datetime, timedelta
    from collections import defaultdict
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    VE_TZ = pytz.timezone('America/Caracas')
    wb = Workbook()
    
    if tipo == 'ventas':
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Incluir COMPLETADO y EN_PROCESO
        ventas = ContratoCliente.objects.filter(estado__in=['COMPLETADO', 'EN_PROCESO'])
        
        # ===== FILTRAR POR ESTADO (NUEVO) =====
        if estado:
            ventas = ventas.filter(estado=estado)
        
        # Convertir a datetime aware
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        # ===== FILTRAR CORRECTAMENTE =====
        # COMPLETADO usa fecha_completado
        # EN_PROCESO usa fecha_creacion
        if fecha_inicio_aware and fecha_fin_aware:
            ventas = ventas.filter(
                Q(estado='COMPLETADO', fecha_completado__gte=fecha_inicio_aware, fecha_completado__lte=fecha_fin_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_inicio_aware, fecha_creacion__lte=fecha_fin_aware)
            )
        elif fecha_inicio_aware:
            ventas = ventas.filter(
                Q(estado='COMPLETADO', fecha_completado__gte=fecha_inicio_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_inicio_aware)
            )
        elif fecha_fin_aware:
            ventas = ventas.filter(
                Q(estado='COMPLETADO', fecha_completado__lte=fecha_fin_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__lte=fecha_fin_aware)
            )
        
        if vendedor:
            ventas = ventas.filter(creado_por_id=vendedor)
        if plan:
            ventas = ventas.filter(plan_contratado_id=plan)
        if busqueda:
            ventas = ventas.filter(
                Q(cliente_potencial__nombre__icontains=busqueda) |
                Q(cliente_potencial__apellido__icontains=busqueda) |
                Q(cliente_potencial__cedula__icontains=busqueda) |
                Q(customer_id__icontains=busqueda)
            )
        
        ventas = ventas.order_by('-fecha_creacion')
        
        # Función para formatear fechas en Venezuela
        def formatear_fecha(fecha):
            if fecha:
                fecha_ve = fecha.astimezone(VE_TZ)
                return fecha_ve.strftime('%d/%m/%Y %H:%M')
            return 'N/A'
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'Plan', 'ODS', 'Fecha', 'Vendedor', 'Estado']
            data = []
            for v in ventas:
                if v.estado == 'COMPLETADO' and v.fecha_completado:
                    fecha_str = formatear_fecha(v.fecha_completado)
                else:
                    fecha_str = formatear_fecha(v.fecha_creacion)
                
                data.append([
                    v.nombre_completo,
                    v.customer_id or 'N/A',
                    v.plan_contratado.nombre if v.plan_contratado else 'N/A',
                    v.ods or 'N/A',
                    fecha_str,
                    v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                    v.get_estado_display()
                ])
        else:
            headers = ['ID', 'Cliente', 'Cédula', 'Teléfono', 'Correo', 'Dirección', 'Plan', 'Fecha', 'Fecha Creación', 'Vendedor', 'Customer ID', 'ODS', 'ATR', 'Estado']
            data = []
            for v in ventas:
                if v.estado == 'COMPLETADO' and v.fecha_completado:
                    fecha_str = formatear_fecha(v.fecha_completado)
                else:
                    fecha_str = formatear_fecha(v.fecha_creacion)
                
                data.append([
                    v.id,
                    v.nombre_completo,
                    v.cedula,
                    v.telefono_principal,
                    v.correo_electronico,
                    v.direccion_detallada[:100] if v.direccion_detallada else 'N/A',
                    v.plan_contratado.nombre if v.plan_contratado else 'N/A',
                    fecha_str,
                    formatear_fecha(v.fecha_creacion),
                    v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                    v.customer_id or 'N/A',
                    v.ods or 'N/A',
                    v.atr or 'N/A',
                    v.get_estado_display()
                ])
        
        ws_resumen = wb.create_sheet("Resumen")
        total_registros = ventas.count()
        completados_count = ventas.filter(estado='COMPLETADO').count()
        en_proceso_count = ventas.filter(estado='EN_PROCESO').count()
        
        resumen_data = [
            ['REPORTE DE VENTAS', ''],
            ['', ''],
            ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ['', ''],
            ['ESTADÍSTICAS:', ''],
            ['Total de contratos:', total_registros],
            ['Completados:', completados_count],
            ['En Proceso:', en_proceso_count],
        ]
        
        # Si hay filtro de estado, mostrarlo
        # Si hay filtro de estado, mostrarlo
        if estado:
            # Obtener el label del estado usando el método get_estado_display()
            # O directamente desde el TextChoices
            try:
                estado_display = ContratoCliente.EstadoContrato(estado).label if hasattr(ContratoCliente, 'EstadoContrato') else estado
            except:
                # Si falla, mostrar el valor directamente
                estado_display = estado
            resumen_data.insert(4, ['Filtro de estado:', estado_display])
                
        for row_idx, row_data in enumerate(resumen_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="FF6B00")
                elif row_idx == 5:
                    cell.font = Font(bold=True)
        
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 30
    
    elif tipo == 'instalaciones':
        ws = wb.active
        ws.title = "Reporte de Instalaciones"
        
        instalaciones = Instalacion.objects.all()
        
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        if fecha_inicio_aware and fecha_fin_aware:
            instalaciones = instalaciones.filter(
                fecha_instalacion__gte=fecha_inicio_aware,
                fecha_instalacion__lte=fecha_fin_aware
            )
        elif fecha_inicio_aware:
            instalaciones = instalaciones.filter(fecha_instalacion__gte=fecha_inicio_aware)
        elif fecha_fin_aware:
            instalaciones = instalaciones.filter(fecha_instalacion__lte=fecha_fin_aware)
        
        if cuadrilla:
            instalaciones = instalaciones.filter(asignacion__cuadrilla_id=cuadrilla)
        if estado == 'completada':
            instalaciones = instalaciones.filter(completada=True)
        elif estado == 'pendiente':
            instalaciones = instalaciones.filter(completada=False)
        if busqueda:
            instalaciones = instalaciones.filter(
                Q(nombre_cliente__icontains=busqueda) |
                Q(cedula_cliente__icontains=busqueda) |
                Q(customer_id__icontains=busqueda)
            )
        
        instalaciones = instalaciones.order_by('-fecha_instalacion')
        
        def formatear_fecha_inst(fecha):
            if fecha:
                fecha_ve = fecha.astimezone(VE_TZ)
                return fecha_ve.strftime('%d/%m/%Y')
            return 'No registrada'
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'ODS', 'Fecha', 'Cuadrilla', 'Estado']
            data = []
            for i in instalaciones:
                data.append([
                    i.nombre_cliente,
                    i.customer_id,
                    i.orden_servicio,
                    formatear_fecha_inst(i.fecha_instalacion),
                    i.asignacion.cuadrilla.nombre if i.asignacion and i.asignacion.cuadrilla else 'N/A',
                    'Completada' if i.completada else 'Pendiente'
                ])
        else:
            headers = ['ID', 'Cliente', 'Cédula', 'Dirección', 'Plan', 'Cuadrilla', 'Fecha', 'Estado', 'Modelo', 'Serial', 'Metros', 'Customer ID', 'ODS']
            data = []
            for i in instalaciones:
                direccion = "N/A"
                try:
                    if i.asignacion and i.asignacion.contrato:
                        direccion = i.asignacion.contrato.direccion_detallada or "N/A"
                except:
                    pass
                
                data.append([
                    i.id,
                    i.nombre_cliente,
                    i.cedula_cliente,
                    direccion[:100] if direccion else 'N/A',
                    i.plan,
                    i.asignacion.cuadrilla.nombre if i.asignacion and i.asignacion.cuadrilla else 'N/A',
                    formatear_fecha_inst(i.fecha_instalacion),
                    'Completada' if i.completada else 'Pendiente',
                    i.modelo_modem.nombre if i.modelo_modem else 'N/A',
                    i.sn_modem or 'N/A',
                    i.metros_utilizados,
                    i.customer_id,
                    i.orden_servicio
                ])
    
    elif tipo == 'soportes':
        ws = wb.active
        ws.title = "Reporte de Soportes"
        
        soportes = Soporte.objects.all()
        
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        if fecha_inicio_aware and fecha_fin_aware:
            soportes = soportes.filter(
                fecha_hora_servicio__gte=fecha_inicio_aware,
                fecha_hora_servicio__lte=fecha_fin_aware
            )
        elif fecha_inicio_aware:
            soportes = soportes.filter(fecha_hora_servicio__gte=fecha_inicio_aware)
        elif fecha_fin_aware:
            soportes = soportes.filter(fecha_hora_servicio__lte=fecha_fin_aware)
        
        if tipo_soporte:
            soportes = soportes.filter(asignacion__ticket__tipo_soporte=tipo_soporte)
        if estado_soporte:
            soportes = soportes.filter(estado=estado_soporte)
        if cuadrilla:
            soportes = soportes.filter(cuadrilla_id=cuadrilla)
        if busqueda:
            soportes = soportes.filter(
                Q(asignacion__ticket__nombre__icontains=busqueda) |
                Q(asignacion__ticket__apellido__icontains=busqueda) |
                Q(asignacion__ticket__cedula__icontains=busqueda)
            )
        
        soportes = soportes.order_by('-fecha_hora_servicio', '-fecha_creacion')
        
        def formatear_fecha_sop(fecha):
            if fecha:
                fecha_ve = fecha.astimezone(VE_TZ)
                return fecha_ve.strftime('%d/%m/%Y')
            return 'N/A'
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'Ticket Padre', 'Fecha', 'Cuadrilla', 'Estado']
            data = []
            for s in soportes:
                try:
                    cliente = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else 'N/A'
                    customer_id = s.asignacion.ticket.customer_id if s.asignacion and s.asignacion.ticket else 'N/A'
                    ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else 'N/A'
                except:
                    cliente = 'N/A'
                    customer_id = 'N/A'
                    ticket_padre = 'N/A'
                
                data.append([
                    cliente,
                    customer_id,
                    ticket_padre,
                    formatear_fecha_sop(s.fecha_hora_servicio or s.fecha_creacion),
                    s.cuadrilla.nombre if s.cuadrilla else 'N/A',
                    s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado
                ])
        else:
            headers = ['ID', 'Ticket Padre', 'Cliente', 'Cédula', 'Tipo', 'Estado', 'Fecha', 'Falla', 'Solución', 'Cuadrilla']
            data = []
            for s in soportes:
                try:
                    ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else 'N/A'
                    cliente = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else 'N/A'
                    cedula = s.asignacion.ticket.cedula if s.asignacion and s.asignacion.ticket else 'N/A'
                    tipo_display = s.asignacion.ticket.get_tipo_soporte_display() if s.asignacion and s.asignacion.ticket else 'N/A'
                except:
                    ticket_padre = 'N/A'
                    cliente = 'N/A'
                    cedula = 'N/A'
                    tipo_display = 'N/A'
                
                data.append([
                    s.id,
                    ticket_padre,
                    cliente,
                    cedula,
                    tipo_display,
                    s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado,
                    formatear_fecha_sop(s.fecha_hora_servicio or s.fecha_creacion),
                    (s.falla_encontrada[:100] + '...') if s.falla_encontrada and len(s.falla_encontrada) > 100 else (s.falla_encontrada or 'N/A'),
                    (s.solucion[:100] + '...') if s.solucion and len(s.solucion) > 100 else (s.solucion or 'N/A'),
                    s.cuadrilla.nombre if s.cuadrilla else 'N/A'
                ])
    
    elif tipo == 'inventario':
        ws = wb.active
        ws.title = "Reporte de Inventario"
        
        inventario = InventarioGlobal.objects.select_related('material')
        
        if material:
            inventario = inventario.filter(material_id=material)
        if busqueda:
            inventario = inventario.filter(material__nombre__icontains=busqueda)
        
        inventario = inventario.order_by('material__nombre')
        
        if reporte_tipo == 'simple':
            headers = ['Material', 'Cantidad', 'Mínimo', 'Estado']
            data = [[
                item.material.nombre,
                item.cantidad,
                item.cantidad_minima,
                'Bajo stock' if item.esta_bajo_stock else 'Normal'
            ] for item in inventario]
        else:
            headers = ['ID', 'Material', 'Cantidad', 'Mínimo', 'Estado', 'Última Actualización', 'Actualizado Por']
            data = [[
                item.id,
                item.material.nombre,
                item.cantidad,
                item.cantidad_minima,
                'Bajo stock' if item.esta_bajo_stock else 'Normal',
                item.ultima_actualizacion.strftime('%d/%m/%Y %H:%M'),
                item.actualizado_por.get_full_name() or item.actualizado_por.username if item.actualizado_por else 'Sistema'
            ] for item in inventario]
    
    elif tipo == 'vendedores':
        import re
        from decimal import Decimal
        
        ws = wb.active
        ws.title = "Reporte de Vendedores"
        
        # Obtener la fecha de referencia para la semana
        if semana_obj:
            fecha_referencia = semana_obj
        else:
            fecha_referencia = datetime.now().date()
        
        # Calcular semana (viernes a jueves)
        dias_desde_viernes = fecha_referencia.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
        fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
        
        # Contratos completados en la semana
        contratos = ContratoCliente.objects.filter(
            estado='COMPLETADO',
            fecha_completado__gte=fecha_inicio_aware,
            fecha_completado__lte=fecha_fin_aware
        )
        
        if vendedor:
            contratos = contratos.filter(creado_por_id=vendedor)
        
        # Obtener vendedores
        vendedores_list = User.objects.filter(
            groups__name__in=['Vendedor', 'Supervisor', 'Administrador']
        ).distinct()
        
        if vendedor:
            vendedores_list = vendedores_list.filter(id=vendedor)
        
        if busqueda:
            vendedores_list = vendedores_list.filter(
                Q(first_name__icontains=busqueda) |
                Q(username__icontains=busqueda) |
                Q(last_name__icontains=busqueda)
            )
        
        # Obtener tasa de cambio
        tasa_obj = TasaCambio.objects.filter(activo=True).first()
        tasa = float(tasa_obj.tasa) if tasa_obj else 0
        
        # Función para calcular comisión por contrato
        def calcular_comision_contrato(plan_nombre, cashea):
            """
            Calcula la comisión por contrato según el plan y si tiene cashea
            
            Planes:
            - 300 Mbps: $8 normal / $12 cashea
            - 400 Mbps: $12 normal / $15 cashea
            - 500 Mbps o más: $15 normal / $17 cashea
            """
            # Extraer número del plan
            numeros = re.findall(r'\d+', plan_nombre)
            if not numeros:
                return 0
            
            velocidad = int(numeros[0])
            
            if velocidad == 300:
                return 12 if cashea else 8
            elif velocidad == 400:
                return 15 if cashea else 12
            elif velocidad >= 500:
                return 17 if cashea else 15
            else:
                return 8 if cashea else 5
        
        data_vendedores = []
        
        for vendedor_obj in vendedores_list:
            contratos_vendedor = contratos.filter(creado_por=vendedor_obj)
            
            total_contratos = 0
            total_comision = 0
            contratos_para_bono = 0  # Solo contratos de 400 Mbps o más
            lista_contratos_detalle = []
            
            for contrato in contratos_vendedor.order_by('-fecha_completado'):
                plan_nombre = contrato.plan_contratado.nombre
                cashea = contrato.cashea
                comision = calcular_comision_contrato(plan_nombre, cashea)
                
                total_contratos += 1
                total_comision += comision
                
                # Verificar si el contrato cuenta para el bono (400 Mbps o más)
                numeros = re.findall(r'\d+', plan_nombre)
                if numeros and int(numeros[0]) >= 400:
                    contratos_para_bono += 1
                
                # Fecha para el detalle
                fecha_ve = contrato.fecha_completado.astimezone(VE_TZ) if contrato.fecha_completado else None
                fecha_str = fecha_ve.strftime('%d/%m/%Y %H:%M') if fecha_ve else 'N/A'
                
                lista_contratos_detalle.append({
                    'cliente': contrato.nombre_completo,
                    'fecha': fecha_str,
                    'plan': plan_nombre,
                    'customer_id': contrato.customer_id or 'N/A',
                    'cashea': 'Sí' if cashea else 'No',
                    'comision': comision
                })
            
            # Calcular bono (solo si tiene 8 o más contratos de 400 Mbps o más)
            bono = 25 if contratos_para_bono >= 8 else 0
            total_con_bono = total_comision + bono
            total_bs = total_con_bono * tasa
            
            # Solo mostrar si tiene contratos o si se está filtrando por vendedor específico
            if total_contratos > 0 or vendedor:
                # Determinar rango para el bono
                if contratos_para_bono >= 8:
                    rango_bono = f"✅ {contratos_para_bono} contratos 400+ Mbps - Bono aplicado"
                else:
                    faltan = 8 - contratos_para_bono
                    rango_bono = f"❌ {contratos_para_bono}/8 contratos 400+ Mbps - Faltan {faltan} para bono"
                
                data_vendedores.append({
                    'vendedor': vendedor_obj.get_full_name() or vendedor_obj.username,
                    'username': vendedor_obj.username,
                    'contratos': total_contratos,
                    'contratos_para_bono': contratos_para_bono,
                    'comision_total': total_comision,
                    'bono': bono,
                    'total_con_bono': total_con_bono,
                    'total_bs': total_bs,
                    'rango_bono': rango_bono,
                    'contratos_detalle': lista_contratos_detalle
                })
        
        # Ordenar por total de contratos (de mayor a menor)
        data_vendedores.sort(key=lambda x: x['contratos'], reverse=True)
        
        # Calcular estadísticas generales
        total_contratos_general = sum(v['contratos'] for v in data_vendedores)
        total_comisiones_general = sum(v['comision_total'] for v in data_vendedores)
        total_bonos_general = sum(v['bono'] for v in data_vendedores)
        total_pagar_usd_general = sum(v['total_con_bono'] for v in data_vendedores)
        total_pagar_bs_general = total_pagar_usd_general * tasa
        vendedores_con_bono = sum(1 for v in data_vendedores if v['bono'] > 0)
        
        tasa_str = f"{tasa:,.2f}" if tasa else "0.00"
        
        # ===== HOJA DE RESUMEN =====
        ws_resumen = wb.create_sheet("Resumen Semanal")
        
        # Títulos y encabezados
        resumen_data = [
            ['REPORTE DE VENDEDORES', ''],
            ['', ''],
            ['Período:', f"{viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"],
            ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ['Tasa de cambio:', f"1 USD = {tasa_str} Bs"],
            ['', ''],
            ['ESTADÍSTICAS GENERALES:', ''],
            ['Total vendedores con ventas:', len(data_vendedores)],
            ['Vendedores que alcanzaron bono:', vendedores_con_bono],
            ['Total contratos en período:', total_contratos_general],
            ['Total comisiones (USD):', f"${total_comisiones_general:,.2f}"],
            ['Total bonos (USD):', f"${total_bonos_general:,.2f}"],
            ['Total a pagar (USD):', f"${total_pagar_usd_general:,.2f}"],
            ['Total a pagar (Bs):', f"Bs {total_pagar_bs_general:,.2f}"],
            ['', ''],
            ['INFORMACIÓN DE COMISIONES:', ''],
            ['Plan 300 Mbps normal:', '$8 USD'],
            ['Plan 300 Mbps con Cashea:', '$12 USD'],
            ['Plan 400 Mbps normal:', '$12 USD'],
            ['Plan 400 Mbps con Cashea:', '$15 USD'],
            ['Plan 500 Mbps o más normal:', '$15 USD'],
            ['Plan 500 Mbps o más con Cashea:', '$17 USD'],
            ['', ''],
            ['BONO:', ''],
            ['Meta:', '8 contratos de 400 Mbps o más'],
            ['Valor del bono:', '$25 USD'],
        ]
        
        # Estilos para la hoja de resumen
        for row_idx, row_data in enumerate(resumen_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="FF6B00")
                elif row_idx == 7 or row_idx == 16 or row_idx == 23:
                    cell.font = Font(bold=True, size=12)
        
        ws_resumen.column_dimensions['A'].width = 35
        ws_resumen.column_dimensions['B'].width = 30
        
        # ===== HOJA DE VENDEDORES (RESUMEN) =====
        headers_vendedores = [
            'Vendedor', 'Username', 'Contratos', 'Contratos 400+ Mbps', 
            'Comisión Total', 'Bono', 'Total USD', 'Total Bs', 'Estado Bono'
        ]
        
        for col_idx, header in enumerate(headers_vendedores, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, v in enumerate(data_vendedores, 2):
            faltan = 8 - v['contratos_para_bono']
            estado_bono = f"✅ Bono alcanzado" if v['bono'] > 0 else f"❌ Faltan {faltan} contratos 400+ Mbps"
            
            ws.cell(row=row_idx, column=1, value=v['vendedor'])
            ws.cell(row=row_idx, column=2, value=v['username'])
            ws.cell(row=row_idx, column=3, value=v['contratos'])
            ws.cell(row=row_idx, column=4, value=f"{v['contratos_para_bono']} / 8")
            ws.cell(row=row_idx, column=5, value=f"${v['comision_total']:,.2f}")
            ws.cell(row=row_idx, column=6, value=f"${v['bono']:,.2f}")
            ws.cell(row=row_idx, column=7, value=f"${v['total_con_bono']:,.2f}")
            ws.cell(row=row_idx, column=8, value=f"Bs {v['total_bs']:,.2f}")
            ws.cell(row=row_idx, column=9, value=estado_bono)
            
            # Alineación centrada para números
            for col in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center')
        
        # Ajustar anchos de columnas
        for col_idx, header in enumerate(headers_vendedores, 1):
            max_length = max(len(header), max([len(str(ws.cell(row=row, column=col_idx).value or '')) for row in range(2, len(data_vendedores) + 2)] or [0]))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
        
        # ===== HOJA DE DETALLE DE CONTRATOS =====
        ws_detalles = wb.create_sheet("Detalle Contratos")
        
        detalles_headers = ['Vendedor', 'Cliente', 'Fecha Completado', 'Plan', 'Cashea', 'Comisión', 'Customer ID']
        
        for col_idx, header in enumerate(detalles_headers, 1):
            cell = ws_detalles.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        detalles_data = []
        for v in data_vendedores:
            for contrato in v['contratos_detalle']:
                detalles_data.append([
                    v['vendedor'],
                    contrato['cliente'],
                    contrato['fecha'],
                    contrato['plan'],
                    contrato['cashea'],
                    f"${contrato['comision']:,.2f}",
                    contrato['customer_id']
                ])
        
        for row_idx, row_data in enumerate(detalles_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_detalles.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar anchos de columnas en detalles
        for col_idx, header in enumerate(detalles_headers, 1):
            max_length = max(len(header), max([len(str(row[col_idx-1])) for row in detalles_data[:100]] or [0]))
            ws_detalles.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 35)
        
        # ===== HOJA DE RESUMEN POR VENDEDOR (DETALLE INDIVIDUAL) =====
        ws_detalle_vendedores = wb.create_sheet("Resumen por Vendedor")
        
        row_idx = 1
        for v in data_vendedores:
            # Título del vendedor
            title_cell = ws_detalle_vendedores.cell(row=row_idx, column=1, value=f"📊 {v['vendedor']} (@{v['username']})")
            title_cell.font = Font(bold=True, size=12, color="FF6B00")
            ws_detalle_vendedores.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            row_idx += 1
            
            # Resumen del vendedor
            resumen_vendedor = [
                ['Total Contratos', f"{v['contratos']}"],
                ['Contratos 400+ Mbps', f"{v['contratos_para_bono']} / 8"],
                ['Comisión Total', f"${v['comision_total']:,.2f}"],
                ['Bono', f"${v['bono']:,.2f}"],
                ['Total a Pagar USD', f"${v['total_con_bono']:,.2f}"],
                ['Total a Pagar Bs', f"Bs {v['total_bs']:,.2f}"],
                ['Estado', v['rango_bono']],
            ]
            
            for r_data in resumen_vendedor:
                ws_detalle_vendedores.cell(row=row_idx, column=1, value=r_data[0])
                ws_detalle_vendedores.cell(row=row_idx, column=2, value=r_data[1])
                row_idx += 1
            
            row_idx += 1
            
            # Tabla de contratos del vendedor
            if v['contratos_detalle']:
                # Encabezados
                headers_contratos = ['#', 'Cliente', 'Fecha', 'Plan', 'Cashea', 'Comisión', 'Customer ID']
                for col, header in enumerate(headers_contratos, 1):
                    cell = ws_detalle_vendedores.cell(row=row_idx, column=col, value=header)
                    cell.fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                row_idx += 1
                
                for idx, contrato in enumerate(v['contratos_detalle'], 1):
                    ws_detalle_vendedores.cell(row=row_idx, column=1, value=idx)
                    ws_detalle_vendedores.cell(row=row_idx, column=2, value=contrato['cliente'])
                    ws_detalle_vendedores.cell(row=row_idx, column=3, value=contrato['fecha'])
                    ws_detalle_vendedores.cell(row=row_idx, column=4, value=contrato['plan'])
                    ws_detalle_vendedores.cell(row=row_idx, column=5, value=contrato['cashea'])
                    ws_detalle_vendedores.cell(row=row_idx, column=6, value=f"${contrato['comision']:,.2f}")
                    ws_detalle_vendedores.cell(row=row_idx, column=7, value=contrato['customer_id'])
                    row_idx += 1
            
            row_idx += 2  # Espacio entre vendedores
        
        # Ajustar anchos de columnas en resumen por vendedor
        ws_detalle_vendedores.column_dimensions['A'].width = 20
        ws_detalle_vendedores.column_dimensions['B'].width = 30
        ws_detalle_vendedores.column_dimensions['C'].width = 20
        ws_detalle_vendedores.column_dimensions['D'].width = 25
        ws_detalle_vendedores.column_dimensions['E'].width = 12
        ws_detalle_vendedores.column_dimensions['F'].width = 15
        ws_detalle_vendedores.column_dimensions['G'].width = 25
    
    else:  # instaladores
        ws = wb.active
        ws.title = "Reporte de Instaladores"
        
        PRECIO_INSTALACION = 15
        PRECIO_CONTRATO = 10
        PRECIOS_SOPORTES = {
            'SOPORTE': 10,
            'RETIRO': 8,
            'MUDANZA': 15,
            'RECABLEADO': 15,
        }
        
        if semana_obj:
            fecha_referencia = semana_obj
        else:
            fecha_referencia = datetime.now().date()
        
        dias_desde_viernes = fecha_referencia.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
        fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
        
        todas_cuadrillas = Cuadrilla.objects.filter(activo=True)
        
        if cuadrilla:
            todas_cuadrillas = todas_cuadrillas.filter(id=cuadrilla)
        
        cuadrillas_dict = defaultdict(lambda: {
            'id': None,
            'cuadrilla': '',
            'instalaciones': 0,
            'monto_instalaciones': 0,
            'soportes': 0,
            'monto_soportes': 0,
            'contratos': 0,
            'monto_contratos': 0,
            'instaladores_set': set(),
            'instaladores_list': []
        })
        
        for cuadrilla_obj in todas_cuadrillas:
            cuadrillas_dict[cuadrilla_obj.nombre]['id'] = cuadrilla_obj.id
            cuadrillas_dict[cuadrilla_obj.nombre]['cuadrilla'] = cuadrilla_obj.nombre
            
            perfiles = cuadrilla_obj.instaladores.all()
            nombres_instaladores = []
            for perfil in perfiles:
                if perfil.usuario:
                    nombres_instaladores.append(perfil.usuario.get_full_name() or perfil.usuario.username)
                    cuadrillas_dict[cuadrilla_obj.nombre]['instaladores_set'].add(perfil.usuario.id)
            cuadrillas_dict[cuadrilla_obj.nombre]['instaladores_list'] = nombres_instaladores
        
        instalaciones = Instalacion.objects.filter(
            completada=True,
            fecha_instalacion__gte=fecha_inicio_aware,
            fecha_instalacion__lte=fecha_fin_aware
        ).select_related('asignacion__cuadrilla')
        
        for inst in instalaciones:
            cuadrilla_obj = inst.asignacion.cuadrilla if inst.asignacion else None
            if not cuadrilla_obj:
                continue
            nombre_cuadrilla = cuadrilla_obj.nombre
            if nombre_cuadrilla in cuadrillas_dict:
                instaladores_hist = inst.instaladores.all()
                
                if instalador:
                    if not instaladores_hist.filter(id=instalador).exists():
                        continue
                
                if busqueda:
                    tiene_coincidencia = False
                    for inst_hist in instaladores_hist:
                        nombre_completo = inst_hist.get_full_name() or inst_hist.username
                        if busqueda.lower() in nombre_completo.lower():
                            tiene_coincidencia = True
                            break
                    if not tiene_coincidencia:
                        continue
                
                cuadrillas_dict[nombre_cuadrilla]['instalaciones'] += 1
                cuadrillas_dict[nombre_cuadrilla]['monto_instalaciones'] += PRECIO_INSTALACION
                
                for inst_hist in instaladores_hist:
                    cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        soportes = Soporte.objects.filter(
            estado='COMPLETADO',
            fecha_creacion__gte=fecha_inicio_aware,
            fecha_creacion__lte=fecha_fin_aware
        ).select_related('cuadrilla')
        
        for sop in soportes:
            if not sop.cuadrilla:
                continue
            nombre_cuadrilla = sop.cuadrilla.nombre
            if nombre_cuadrilla in cuadrillas_dict:
                instaladores_hist = sop.instaladores.all()
                
                if instalador:
                    if not instaladores_hist.filter(id=instalador).exists():
                        continue
                
                if busqueda:
                    tiene_coincidencia = False
                    for inst_hist in instaladores_hist:
                        nombre_completo = inst_hist.get_full_name() or inst_hist.username
                        if busqueda.lower() in nombre_completo.lower():
                            tiene_coincidencia = True
                            break
                    if not tiene_coincidencia:
                        continue
                
                try:
                    tipo = sop.asignacion.ticket.tipo_soporte if sop.asignacion and sop.asignacion.ticket else 'SOPORTE'
                    precio = PRECIOS_SOPORTES.get(tipo, 10)
                except:
                    precio = 10
                
                cuadrillas_dict[nombre_cuadrilla]['soportes'] += 1
                cuadrillas_dict[nombre_cuadrilla]['monto_soportes'] += precio
                
                for inst_hist in instaladores_hist:
                    cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        instaladores_users = User.objects.filter(groups__name='Instalador')
        
        if instalador:
            instaladores_users = instaladores_users.filter(id=instalador)
        
        for instalador_user in instaladores_users:
            perfil = PerfilUsuario.objects.filter(usuario=instalador_user).first()
            if not perfil:
                continue
            cuadrillas_del_instalador = perfil.cuadrillas.all()
            if not cuadrillas_del_instalador.exists():
                continue
            contratos_instalador = ContratoCliente.objects.filter(
                estado='COMPLETADO',
                creado_por=instalador_user,
                fecha_completado__gte=fecha_inicio_aware,
                fecha_completado__lte=fecha_fin_aware
            )
            for contrato in contratos_instalador:
                for cuadrilla_inst in cuadrillas_del_instalador:
                    nombre_cuadrilla = cuadrilla_inst.nombre
                    if nombre_cuadrilla in cuadrillas_dict:
                        cuadrillas_dict[nombre_cuadrilla]['contratos'] += 1
                        cuadrillas_dict[nombre_cuadrilla]['monto_contratos'] += PRECIO_CONTRATO
                        cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(instalador_user.id)
        
        tasa_obj = TasaCambio.objects.filter(activo=True).first()
        tasa = float(tasa_obj.tasa) if tasa_obj else 0
        
        data_cuadrillas = []
        for nombre, data in cuadrillas_dict.items():
            if data['instalaciones'] > 0 or data['soportes'] > 0 or data['contratos'] > 0 or cuadrilla:
                total_usd = data['monto_instalaciones'] + data['monto_soportes'] + data['monto_contratos']
                total_bs = total_usd * tasa
                
                nombres_instaladores_final = []
                for inst_id in data['instaladores_set']:
                    try:
                        inst = User.objects.get(id=inst_id)
                        nombres_instaladores_final.append(inst.get_full_name() or inst.username)
                    except:
                        pass
                
                data_cuadrillas.append({
                    'cuadrilla': data['cuadrilla'],
                    'instalaciones': data['instalaciones'],
                    'monto_instalaciones': data['monto_instalaciones'],
                    'soportes': data['soportes'],
                    'monto_soportes': data['monto_soportes'],
                    'contratos': data['contratos'],
                    'monto_contratos': data['monto_contratos'],
                    'total_usd': total_usd,
                    'total_bs': total_bs,
                    'instaladores_list': nombres_instaladores_final or data['instaladores_list']
                })
        
        data_cuadrillas.sort(key=lambda x: x['instalaciones'] + x['soportes'] + x['contratos'], reverse=True)
        
        headers = ['Cuadrilla', 'Instalaciones', 'Monto Instalaciones', 'Soportes', 'Monto Soportes', 'Contratos', 'Monto Contratos', 'Total USD', 'Total Bs', 'Instaladores']
        data = [[
            item['cuadrilla'],
            item['instalaciones'],
            f"${item['monto_instalaciones']}",
            item['soportes'],
            f"${item['monto_soportes']}",
            item['contratos'],
            f"${item['monto_contratos']}",
            f"${item['total_usd']}",
            f"Bs {item['total_bs']:,.2f}",
            ', '.join(item['instaladores_list']) if item['instaladores_list'] else 'Sin instaladores'
        ] for item in data_cuadrillas]
    
    # Aplicar estilos generales
    if tipo not in ['vendedores', 'instaladores']:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            max_length = max(len(str(headers[col-1])), max([len(str(row[col-1])) for row in data[:100]] or [0]))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    else:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            max_length = max(len(str(headers[col-1])), max([len(str(row[col-1])) for row in data[:100]] or [0]))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)
    
    filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


def exportar_pdf(request, tipo, reporte_tipo, fecha_desde_obj, fecha_hasta_obj,
                 busqueda, vendedor, plan, cuadrilla, estado,
                 tipo_soporte, estado_soporte, material, semana_obj, instalador):
    """Exportar datos a PDF con filtros usando datetime aware"""
    
    import io
    import pytz
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    VE_TZ = pytz.timezone('America/Caracas')
    
    def formatear_fecha_pdf(fecha):
        if fecha:
            fecha_ve = fecha.astimezone(VE_TZ)
            return fecha_ve.strftime('%d/%m/%Y %H:%M')
        return 'N/A'
    
    if tipo == 'ventas':
        titulo = "Reporte de Ventas"
        
        datos = ContratoCliente.objects.filter(estado__in=['COMPLETADO', 'EN_PROCESO'])
        
        # ===== FILTRAR POR ESTADO (NUEVO) =====
        if estado:
            datos = datos.filter(estado=estado)
        
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        # ===== FILTRAR CORRECTAMENTE =====
        if fecha_inicio_aware and fecha_fin_aware:
            datos = datos.filter(
                Q(estado='COMPLETADO', fecha_completado__gte=fecha_inicio_aware, fecha_completado__lte=fecha_fin_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_inicio_aware, fecha_creacion__lte=fecha_fin_aware)
            )
        elif fecha_inicio_aware:
            datos = datos.filter(
                Q(estado='COMPLETADO', fecha_completado__gte=fecha_inicio_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__gte=fecha_inicio_aware)
            )
        elif fecha_fin_aware:
            datos = datos.filter(
                Q(estado='COMPLETADO', fecha_completado__lte=fecha_fin_aware) |
                Q(estado='EN_PROCESO', fecha_creacion__lte=fecha_fin_aware)
            )
        
        if vendedor:
            datos = datos.filter(creado_por_id=vendedor)
        if plan:
            datos = datos.filter(plan_contratado_id=plan)
        if busqueda:
            datos = datos.filter(
                Q(cliente_potencial__nombre__icontains=busqueda) |
                Q(cliente_potencial__apellido__icontains=busqueda) |
                Q(cliente_potencial__cedula__icontains=busqueda) |
                Q(customer_id__icontains=busqueda)
            )
        
        datos = datos.order_by('-fecha_creacion')
        total_registros = datos.count()
        completados = datos.filter(estado='COMPLETADO').count()
        en_proceso = datos.filter(estado='EN_PROCESO').count()
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'Plan', 'ODS', 'Fecha', 'Vendedor', 'Estado']
            rows = []
            for v in datos:
                if v.estado == 'COMPLETADO' and v.fecha_completado:
                    fecha_str = formatear_fecha_pdf(v.fecha_completado)
                else:
                    fecha_str = formatear_fecha_pdf(v.fecha_creacion)
                
                rows.append([
                    v.nombre_completo,
                    v.customer_id or 'N/A',
                    v.plan_contratado.nombre if v.plan_contratado else 'N/A',
                    v.ods or 'N/A',
                    fecha_str,
                    v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                    v.get_estado_display()
                ])
        else:
            headers = ['ID', 'Cliente', 'Cédula', 'Teléfono', 'Plan', 'Fecha', 'Fecha Creación', 'Vendedor', 'Customer ID', 'ODS', 'Estado']
            rows = []
            for v in datos:
                if v.estado == 'COMPLETADO' and v.fecha_completado:
                    fecha_str = formatear_fecha_pdf(v.fecha_completado)
                else:
                    fecha_str = formatear_fecha_pdf(v.fecha_creacion)
                
                rows.append([
                    str(v.id),
                    v.nombre_completo,
                    v.cedula,
                    v.telefono_principal,
                    v.plan_contratado.nombre if v.plan_contratado else 'N/A',
                    fecha_str,
                    formatear_fecha_pdf(v.fecha_creacion),
                    v.creado_por.get_full_name() or v.creado_por.username if v.creado_por else 'N/A',
                    v.customer_id or 'N/A',
                    v.ods or 'N/A',
                    v.get_estado_display()
                ])
    
    elif tipo == 'instalaciones':
        titulo = "Reporte de Instalaciones"
        
        datos = Instalacion.objects.all()
        
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        if fecha_inicio_aware and fecha_fin_aware:
            datos = datos.filter(
                fecha_instalacion__gte=fecha_inicio_aware,
                fecha_instalacion__lte=fecha_fin_aware
            )
        elif fecha_inicio_aware:
            datos = datos.filter(fecha_instalacion__gte=fecha_inicio_aware)
        elif fecha_fin_aware:
            datos = datos.filter(fecha_instalacion__lte=fecha_fin_aware)
        
        if cuadrilla:
            datos = datos.filter(asignacion__cuadrilla_id=cuadrilla)
        if estado == 'completada':
            datos = datos.filter(completada=True)
        elif estado == 'pendiente':
            datos = datos.filter(completada=False)
        if busqueda:
            datos = datos.filter(
                Q(nombre_cliente__icontains=busqueda) |
                Q(cedula_cliente__icontains=busqueda) |
                Q(customer_id__icontains=busqueda)
            )
        
        datos = datos.order_by('-fecha_instalacion')
        total_registros = datos.count()
        completadas = datos.filter(completada=True).count()
        pendientes = datos.filter(completada=False).count()
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'ODS', 'Fecha', 'Cuadrilla', 'Estado']
            rows = []
            for i in datos:
                rows.append([
                    i.nombre_cliente,
                    i.customer_id,
                    i.orden_servicio,
                    formatear_fecha_pdf(i.fecha_instalacion) if i.fecha_instalacion else 'No registrada',
                    i.asignacion.cuadrilla.nombre if i.asignacion and i.asignacion.cuadrilla else 'N/A',
                    'Completada' if i.completada else 'Pendiente'
                ])
        else:
            headers = ['ID', 'Cliente', 'Cédula', 'Plan', 'Cuadrilla', 'Fecha', 'Estado', 'Modelo', 'Serial', 'Metros']
            rows = []
            for i in datos:
                rows.append([
                    str(i.id),
                    i.nombre_cliente,
                    i.cedula_cliente,
                    i.plan,
                    i.asignacion.cuadrilla.nombre if i.asignacion and i.asignacion.cuadrilla else 'N/A',
                    formatear_fecha_pdf(i.fecha_instalacion) if i.fecha_instalacion else 'N/A',
                    'Completada' if i.completada else 'Pendiente',
                    i.modelo_modem.nombre if i.modelo_modem else 'N/A',
                    i.sn_modem or 'N/A',
                    str(i.metros_utilizados)
                ])
    
    elif tipo == 'soportes':
        titulo = "Reporte de Soportes"
        
        datos = Soporte.objects.all()
        
        fecha_inicio_aware = None
        fecha_fin_aware = None
        
        if fecha_desde_obj:
            fecha_inicio_aware = VE_TZ.localize(datetime.combine(fecha_desde_obj, datetime.min.time()))
        
        if fecha_hasta_obj:
            fecha_fin_aware = VE_TZ.localize(datetime.combine(fecha_hasta_obj, datetime.max.time()))
        
        if fecha_inicio_aware and fecha_fin_aware:
            datos = datos.filter(
                fecha_hora_servicio__gte=fecha_inicio_aware,
                fecha_hora_servicio__lte=fecha_fin_aware
            )
        elif fecha_inicio_aware:
            datos = datos.filter(fecha_hora_servicio__gte=fecha_inicio_aware)
        elif fecha_fin_aware:
            datos = datos.filter(fecha_hora_servicio__lte=fecha_fin_aware)
        
        if tipo_soporte:
            datos = datos.filter(asignacion__ticket__tipo_soporte=tipo_soporte)
        if estado_soporte:
            datos = datos.filter(estado=estado_soporte)
        if cuadrilla:
            datos = datos.filter(cuadrilla_id=cuadrilla)
        if busqueda:
            datos = datos.filter(
                Q(asignacion__ticket__nombre__icontains=busqueda) |
                Q(asignacion__ticket__apellido__icontains=busqueda) |
                Q(asignacion__ticket__cedula__icontains=busqueda)
            )
        
        datos = datos.order_by('-fecha_hora_servicio', '-fecha_creacion')
        total_registros = datos.count()
        completados = datos.filter(estado='COMPLETADO').count()
        pendientes = datos.filter(estado='PENDIENTE').count()
        en_proceso = datos.filter(estado='EN_PROCESO').count()
        
        if reporte_tipo == 'simple':
            headers = ['Cliente', 'Customer ID', 'Ticket Padre', 'Fecha', 'Cuadrilla', 'Estado']
            rows = []
            for s in datos:
                try:
                    cliente = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else 'N/A'
                    customer_id = s.asignacion.ticket.customer_id if s.asignacion and s.asignacion.ticket else 'N/A'
                    ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else 'N/A'
                except:
                    cliente = 'N/A'
                    customer_id = 'N/A'
                    ticket_padre = 'N/A'
                
                rows.append([
                    cliente,
                    customer_id,
                    ticket_padre,
                    formatear_fecha_pdf(s.fecha_hora_servicio or s.fecha_creacion),
                    s.cuadrilla.nombre if s.cuadrilla else 'N/A',
                    s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado
                ])
        else:
            headers = ['ID', 'Ticket Padre', 'Cliente', 'Cédula', 'Tipo', 'Estado', 'Fecha', 'Falla', 'Solución']
            rows = []
            for s in datos:
                try:
                    ticket_padre = s.asignacion.ticket.ticket_padre if s.asignacion and s.asignacion.ticket else 'N/A'
                    cliente = s.asignacion.ticket.nombre_completo if s.asignacion and s.asignacion.ticket else 'N/A'
                    cedula = s.asignacion.ticket.cedula if s.asignacion and s.asignacion.ticket else 'N/A'
                    tipo_display = s.asignacion.ticket.get_tipo_soporte_display() if s.asignacion and s.asignacion.ticket else 'N/A'
                except:
                    ticket_padre = 'N/A'
                    cliente = 'N/A'
                    cedula = 'N/A'
                    tipo_display = 'N/A'
                
                rows.append([
                    str(s.id),
                    ticket_padre,
                    cliente,
                    cedula,
                    tipo_display,
                    s.get_estado_display() if hasattr(s, 'get_estado_display') else s.estado,
                    formatear_fecha_pdf(s.fecha_hora_servicio or s.fecha_creacion),
                    (s.falla_encontrada[:80] + '...') if s.falla_encontrada and len(s.falla_encontrada) > 80 else (s.falla_encontrada or 'N/A'),
                    (s.solucion[:80] + '...') if s.solucion and len(s.solucion) > 80 else (s.solucion or 'N/A')
                ])
    
    elif tipo == 'inventario':
        titulo = "Reporte de Inventario"
        
        datos = InventarioGlobal.objects.select_related('material')
        
        if material:
            datos = datos.filter(material_id=material)
        if busqueda:
            datos = datos.filter(material__nombre__icontains=busqueda)
        
        datos = datos.order_by('material__nombre')
        total_registros = datos.count()
        total_unidades = sum(item.cantidad for item in datos)
        bajo_stock = sum(1 for item in datos if item.esta_bajo_stock)
        
        if reporte_tipo == 'simple':
            headers = ['Material', 'Cantidad', 'Mínimo', 'Estado']
            rows = [[
                item.material.nombre,
                str(item.cantidad),
                str(item.cantidad_minima),
                'Bajo stock' if item.esta_bajo_stock else 'Normal'
            ] for item in datos]
        else:
            headers = ['ID', 'Material', 'Cantidad', 'Mínimo', 'Estado', 'Última Actualización', 'Actualizado Por']
            rows = [[
                str(item.id),
                item.material.nombre,
                str(item.cantidad),
                str(item.cantidad_minima),
                'Bajo stock' if item.esta_bajo_stock else 'Normal',
                item.ultima_actualizacion.strftime('%d/%m/%Y %H:%M'),
                item.actualizado_por.get_full_name() or item.actualizado_por.username if item.actualizado_por else 'Sistema'
            ] for item in datos]
    
    elif tipo == 'vendedores':
        titulo = "Reporte de Vendedores"
        
        if semana_obj:
            fecha_referencia = semana_obj
        else:
            fecha_referencia = datetime.now().date()
        
        dias_desde_viernes = fecha_referencia.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
        fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
        
        contratos = ContratoCliente.objects.filter(
            estado='COMPLETADO',
            fecha_completado__gte=fecha_inicio_aware,
            fecha_completado__lte=fecha_fin_aware
        )
        
        if vendedor:
            contratos = contratos.filter(creado_por_id=vendedor)
        
        vendedores_list = User.objects.filter(
            groups__name__in=['Vendedor', 'Supervisor', 'Administrador']
        ).distinct()
        
        if vendedor:
            vendedores_list = vendedores_list.filter(id=vendedor)
        
        if busqueda:
            vendedores_list = vendedores_list.filter(
                Q(first_name__icontains=busqueda) |
                Q(username__icontains=busqueda) |
                Q(last_name__icontains=busqueda)
            )
        
        tasa_obj = TasaCambio.objects.filter(activo=True).first()
        tasa = float(tasa_obj.tasa) if tasa_obj else 0
        
        # Función para calcular comisión por contrato
        def calcular_comision_contrato_pdf(plan_nombre, cashea):
            import re
            numeros = re.findall(r'\d+', plan_nombre)
            if not numeros:
                return 0
            velocidad = int(numeros[0])
            
            if velocidad == 300:
                return 12 if cashea else 8
            elif velocidad == 400:
                return 15 if cashea else 12
            elif velocidad >= 500:
                return 17 if cashea else 15
            else:
                return 8 if cashea else 5
        
        data_vendedores = []
        
        for vendedor_obj in vendedores_list:
            contratos_vendedor = contratos.filter(creado_por=vendedor_obj)
            
            total_contratos = 0
            total_comision = 0
            contratos_para_bono = 0  # Solo contratos de 400 Mbps o más
            contratos_detalle = []
            
            for contrato in contratos_vendedor.order_by('-fecha_completado'):
                plan_nombre = contrato.plan_contratado.nombre
                cashea = contrato.cashea
                comision = calcular_comision_contrato_pdf(plan_nombre, cashea)
                
                total_contratos += 1
                total_comision += comision
                
                # Verificar si el contrato cuenta para el bono (400 Mbps o más)
                import re
                numeros = re.findall(r'\d+', plan_nombre)
                if numeros and int(numeros[0]) >= 400:
                    contratos_para_bono += 1
                
                fecha_ve = contrato.fecha_completado.astimezone(VE_TZ) if contrato.fecha_completado else None
                fecha_str = fecha_ve.strftime('%d/%m/%Y') if fecha_ve else 'N/A'
                
                contratos_detalle.append({
                    'cliente': contrato.nombre_completo,
                    'fecha': fecha_str,
                    'plan': plan_nombre,
                    'cashea': 'Sí' if cashea else 'No',
                    'comision': comision
                })
            
            # Calcular bono (solo si tiene 8 o más contratos de 400 Mbps o más)
            bono = 25 if contratos_para_bono >= 8 else 0
            total_con_bono = total_comision + bono
            total_bs = total_con_bono * tasa
            
            if total_contratos > 0 or vendedor:
                if contratos_para_bono >= 8:
                    rango_bono = f"✅ {contratos_para_bono} contratos 400+ Mbps - Bono aplicado"
                else:
                    faltan = 8 - contratos_para_bono
                    rango_bono = f"❌ {contratos_para_bono}/8 contratos 400+ Mbps - Faltan {faltan} para bono"
                
                data_vendedores.append({
                    'vendedor': vendedor_obj.get_full_name() or vendedor_obj.username,
                    'username': vendedor_obj.username,
                    'contratos': total_contratos,
                    'contratos_para_bono': contratos_para_bono,
                    'comision_total': total_comision,
                    'bono': bono,
                    'total_con_bono': total_con_bono,
                    'total_bs': total_bs,
                    'rango_bono': rango_bono,
                    'contratos_detalle': contratos_detalle
                })
        
        data_vendedores.sort(key=lambda x: x['contratos'], reverse=True)
        
        total_contratos_general = sum(v['contratos'] for v in data_vendedores)
        total_comisiones_general = sum(v['comision_total'] for v in data_vendedores)
        total_bonos_general = sum(v['bono'] for v in data_vendedores)
        total_pagar_usd_general = sum(v['total_con_bono'] for v in data_vendedores)
        total_pagar_bs_general = total_pagar_usd_general * tasa
        vendedores_con_bono = sum(1 for v in data_vendedores if v['bono'] > 0)
        
        # Preparar datos para la tabla
        rows = [[
            v['vendedor'],
            v['username'],
            str(v['contratos']),
            f"{v['contratos_para_bono']} / 8",
            f"${v['comision_total']:,.2f}",
            f"${v['bono']:,.2f}",
            f"${v['total_con_bono']:,.2f}",
            f"Bs {v['total_bs']:,.2f}",
            v['rango_bono']
        ] for v in data_vendedores]
        
        headers = ['Vendedor', 'Usuario', 'Contratos', '400+ Mbps', 'Comisión Total', 'Bono', 'Total USD', 'Total Bs', 'Estado Bono']
        
        total_registros = len(rows)
        completados = total_contratos_general
        en_proceso = 0
    
    else:  # instaladores
        titulo = "Reporte de Instaladores"
        
        PRECIO_INSTALACION = 15
        PRECIO_CONTRATO = 10
        PRECIOS_SOPORTES = {
            'SOPORTE': 10,
            'RETIRO': 8,
            'MUDANZA': 15,
            'RECABLEADO': 15,
        }
        
        if semana_obj:
            fecha_referencia = semana_obj
        else:
            fecha_referencia = datetime.now().date()
        
        dias_desde_viernes = fecha_referencia.weekday() - 4
        if dias_desde_viernes < 0:
            dias_desde_viernes += 7
        
        viernes_inicio = fecha_referencia - timedelta(days=dias_desde_viernes)
        jueves_fin = viernes_inicio + timedelta(days=6)
        
        fecha_inicio_aware = VE_TZ.localize(datetime.combine(viernes_inicio, datetime.min.time()))
        fecha_fin_aware = VE_TZ.localize(datetime.combine(jueves_fin, datetime.max.time()))
        
        todas_cuadrillas = Cuadrilla.objects.filter(activo=True)
        
        if cuadrilla:
            todas_cuadrillas = todas_cuadrillas.filter(id=cuadrilla)
        
        cuadrillas_dict = defaultdict(lambda: {
            'cuadrilla': '',
            'instalaciones': 0,
            'monto_instalaciones': 0,
            'soportes': 0,
            'monto_soportes': 0,
            'contratos': 0,
            'monto_contratos': 0,
            'instaladores_set': set(),
            'instaladores_list': []
        })
        
        for cuadrilla_obj in todas_cuadrillas:
            cuadrillas_dict[cuadrilla_obj.nombre]['cuadrilla'] = cuadrilla_obj.nombre
            
            perfiles = cuadrilla_obj.instaladores.all()
            nombres_instaladores = []
            for perfil in perfiles:
                if perfil.usuario:
                    nombres_instaladores.append(perfil.usuario.get_full_name() or perfil.usuario.username)
                    cuadrillas_dict[cuadrilla_obj.nombre]['instaladores_set'].add(perfil.usuario.id)
            cuadrillas_dict[cuadrilla_obj.nombre]['instaladores_list'] = nombres_instaladores
        
        instalaciones = Instalacion.objects.filter(
            completada=True,
            fecha_instalacion__gte=fecha_inicio_aware,
            fecha_instalacion__lte=fecha_fin_aware
        ).select_related('asignacion__cuadrilla')
        
        for inst in instalaciones:
            cuadrilla_obj = inst.asignacion.cuadrilla if inst.asignacion else None
            if not cuadrilla_obj:
                continue
            nombre_cuadrilla = cuadrilla_obj.nombre
            if nombre_cuadrilla in cuadrillas_dict:
                instaladores_hist = inst.instaladores.all()
                
                if instalador:
                    if not instaladores_hist.filter(id=instalador).exists():
                        continue
                
                if busqueda:
                    tiene_coincidencia = False
                    for inst_hist in instaladores_hist:
                        nombre_completo = inst_hist.get_full_name() or inst_hist.username
                        if busqueda.lower() in nombre_completo.lower():
                            tiene_coincidencia = True
                            break
                    if not tiene_coincidencia:
                        continue
                
                cuadrillas_dict[nombre_cuadrilla]['instalaciones'] += 1
                cuadrillas_dict[nombre_cuadrilla]['monto_instalaciones'] += PRECIO_INSTALACION
                
                for inst_hist in instaladores_hist:
                    cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        soportes = Soporte.objects.filter(
            estado='COMPLETADO',
            fecha_creacion__gte=fecha_inicio_aware,
            fecha_creacion__lte=fecha_fin_aware
        ).select_related('cuadrilla')
        
        for sop in soportes:
            if not sop.cuadrilla:
                continue
            nombre_cuadrilla = sop.cuadrilla.nombre
            if nombre_cuadrilla in cuadrillas_dict:
                instaladores_hist = sop.instaladores.all()
                
                if instalador:
                    if not instaladores_hist.filter(id=instalador).exists():
                        continue
                
                if busqueda:
                    tiene_coincidencia = False
                    for inst_hist in instaladores_hist:
                        nombre_completo = inst_hist.get_full_name() or inst_hist.username
                        if busqueda.lower() in nombre_completo.lower():
                            tiene_coincidencia = True
                            break
                    if not tiene_coincidencia:
                        continue
                
                try:
                    tipo = sop.asignacion.ticket.tipo_soporte if sop.asignacion and sop.asignacion.ticket else 'SOPORTE'
                    precio = PRECIOS_SOPORTES.get(tipo, 10)
                except:
                    precio = 10
                
                cuadrillas_dict[nombre_cuadrilla]['soportes'] += 1
                cuadrillas_dict[nombre_cuadrilla]['monto_soportes'] += precio
                
                for inst_hist in instaladores_hist:
                    cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(inst_hist.id)
        
        instaladores_users = User.objects.filter(groups__name='Instalador')
        
        if instalador:
            instaladores_users = instaladores_users.filter(id=instalador)
        
        for instalador_user in instaladores_users:
            perfil = PerfilUsuario.objects.filter(usuario=instalador_user).first()
            if not perfil:
                continue
            cuadrillas_del_instalador = perfil.cuadrillas.all()
            if not cuadrillas_del_instalador.exists():
                continue
            contratos_instalador = ContratoCliente.objects.filter(
                estado='COMPLETADO',
                creado_por=instalador_user,
                fecha_completado__gte=fecha_inicio_aware,
                fecha_completado__lte=fecha_fin_aware
            )
            for contrato in contratos_instalador:
                for cuadrilla_inst in cuadrillas_del_instalador:
                    nombre_cuadrilla = cuadrilla_inst.nombre
                    if nombre_cuadrilla in cuadrillas_dict:
                        cuadrillas_dict[nombre_cuadrilla]['contratos'] += 1
                        cuadrillas_dict[nombre_cuadrilla]['monto_contratos'] += PRECIO_CONTRATO
                        cuadrillas_dict[nombre_cuadrilla]['instaladores_set'].add(instalador_user.id)
        
        tasa_obj = TasaCambio.objects.filter(activo=True).first()
        tasa = float(tasa_obj.tasa) if tasa_obj else 0
        
        data_cuadrillas = []
        for nombre, data in cuadrillas_dict.items():
            if data['instalaciones'] > 0 or data['soportes'] > 0 or data['contratos'] > 0 or cuadrilla:
                total_usd = data['monto_instalaciones'] + data['monto_soportes'] + data['monto_contratos']
                total_bs = total_usd * tasa
                
                nombres_instaladores_final = []
                for inst_id in data['instaladores_set']:
                    try:
                        inst = User.objects.get(id=inst_id)
                        nombres_instaladores_final.append(inst.get_full_name() or inst.username)
                    except:
                        pass
                
                data_cuadrillas.append({
                    'cuadrilla': data['cuadrilla'],
                    'instalaciones': data['instalaciones'],
                    'monto_instalaciones': data['monto_instalaciones'],
                    'soportes': data['soportes'],
                    'monto_soportes': data['monto_soportes'],
                    'contratos': data['contratos'],
                    'monto_contratos': data['monto_contratos'],
                    'total_usd': total_usd,
                    'total_bs': total_bs,
                    'instaladores_list': nombres_instaladores_final or data['instaladores_list']
                })
        
        data_cuadrillas.sort(key=lambda x: x['instalaciones'] + x['soportes'] + x['contratos'], reverse=True)
        
        rows = [[
            item['cuadrilla'],
            item['instalaciones'],
            f"${item['monto_instalaciones']}",
            item['soportes'],
            f"${item['monto_soportes']}",
            item['contratos'],
            f"${item['monto_contratos']}",
            f"${item['total_usd']}",
            f"Bs {item['total_bs']:,.2f}",
            ', '.join(item['instaladores_list']) if item['instaladores_list'] else 'Sin instaladores'
        ] for item in data_cuadrillas]
        
        headers = ['Cuadrilla', 'Instalaciones', 'Monto Instalaciones', 'Soportes', 'Monto Soportes', 'Contratos', 'Monto Contratos', 'Total USD', 'Total Bs', 'Instaladores']
        
        total_registros = len(rows)
    
    # Construcción del PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, textColor=colors.HexColor('#FF6B00'))
    title = Paragraph(titulo, title_style)
    
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
    fecha_texto = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    if tipo == 'vendedores':
        fecha_texto += f" | Período: {viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"
        fecha_texto += f" | Tasa: 1 USD = {tasa:,.2f} Bs"
    elif tipo == 'instaladores':
        fecha_texto += f" | Período: {viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"
        fecha_texto += f" | Tasa: 1 USD = {tasa:,.2f} Bs"
    elif fecha_desde_obj or fecha_hasta_obj:
        desde_str = fecha_desde_obj.strftime('%d/%m/%Y') if fecha_desde_obj else 'inicio'
        hasta_str = fecha_hasta_obj.strftime('%d/%m/%Y') if fecha_hasta_obj else 'actual'
        fecha_texto += f" | Periodo: {desde_str} al {hasta_str}"
    
    fecha_paragraph = Paragraph(fecha_texto, date_style)
    
    stats_data = [['ESTADÍSTICAS', '']]
    
    if tipo == 'ventas':
        stats_data.extend([
            ['Total registros:', str(len(rows))],
            ['Completados:', str(completados)],
            ['En Proceso:', str(en_proceso)],
        ])
        # Si hay filtro de estado, mostrarlo
        # Si hay filtro de estado, mostrarlo
        if estado:
            # Obtener el label del estado usando el TextChoices
            try:
                estado_display = ContratoCliente.EstadoContrato(estado).label if hasattr(ContratoCliente, 'EstadoContrato') else estado
            except:
                # Si falla, mostrar el valor directamente
                estado_display = estado
            stats_data.insert(1, ['Filtro de estado:', estado_display])
    elif tipo == 'instalaciones':
        stats_data.extend([
            ['Total instalaciones:', str(len(rows))],
            ['Completadas:', str(completadas)],
            ['Pendientes:', str(pendientes)],
        ])
    elif tipo == 'soportes':
        stats_data.extend([
            ['Total soportes:', str(len(rows))],
            ['Completados:', str(completados)],
            ['En proceso:', str(en_proceso)],
            ['Pendientes:', str(pendientes)],
        ])
    elif tipo == 'inventario':
        stats_data.extend([
            ['Total materiales:', str(total_registros)],
            ['Total unidades:', str(total_unidades)],
            ['Bajo stock:', str(bajo_stock)],
        ])
    elif tipo == 'vendedores':
        stats_data.extend([
            ['Período:', f"{viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"],
            ['Total vendedores con ventas:', str(len(rows))],
            ['Total contratos en período:', str(total_contratos_general)],
            ['Total a pagar (USD):', f"${total_pagar_usd_general:,.2f}"],
            ['Total a pagar (Bs):', f"Bs {total_pagar_bs_general:,.2f}"],
            ['Tasa de cambio:', f"1 USD = {tasa:,.2f} Bs"],
        ])
    else:  # instaladores
        total_instalaciones = sum(r[1] for r in rows if isinstance(r[1], (int, float)))
        total_soportes = sum(r[3] for r in rows if isinstance(r[3], (int, float)))
        total_contratos = sum(r[5] for r in rows if isinstance(r[5], (int, float)))
        total_pagar_usd = sum(float(r[7].replace('$', '').replace(',', '')) for r in rows)
        total_pagar_bs = sum(float(r[8].replace('Bs ', '').replace('Bs', '').replace(',', '').strip()) for r in rows)
        
        stats_data.extend([
            ['Período:', f"{viernes_inicio.strftime('%d/%m/%Y')} - {jueves_fin.strftime('%d/%m/%Y')}"],
            ['Total cuadrillas:', str(len(rows))],
            ['Instalaciones:', str(total_instalaciones)],
            ['Soportes:', str(total_soportes)],
            ['Contratos:', str(total_contratos)],
            ['Total a pagar (USD):', f"${total_pagar_usd:,.2f}"],
            ['Total a pagar (Bs):', f"Bs {total_pagar_bs:,.2f}"],
            ['Tasa de cambio:', f"1 USD = {tasa:,.2f} Bs"],
            ['Precio instalación:', f"${PRECIO_INSTALACION}"],
            ['Precio contrato:', f"${PRECIO_CONTRATO}"],
            ['Precios soportes:', 'Soporte: $10, Retiro: $8, Mudanza: $15, Recableado: $15'],
        ])
    
    stats_table = Table(stats_data, colWidths=[150, 200])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements = [title, Spacer(1, 12), fecha_paragraph, Spacer(1, 12), stats_table, Spacer(1, 20)]
    
    if rows:
        max_rows_per_page = 15
        num_pages = (len(rows) + max_rows_per_page - 1) // max_rows_per_page
        
        for page in range(num_pages):
            start_idx = page * max_rows_per_page
            end_idx = min(start_idx + max_rows_per_page, len(rows))
            page_rows = rows[start_idx:end_idx]
            
            table_data = [headers] + page_rows
            table = Table(table_data, repeatRows=1)
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B00')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            table.setStyle(TableStyle(table_style))
            elements.append(table)
            
            if page < num_pages - 1:
                elements.append(PageBreak())
    else:
        no_data = Paragraph("No hay datos disponibles para los filtros seleccionados", styles['Normal'])
        elements.append(no_data)
    
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response